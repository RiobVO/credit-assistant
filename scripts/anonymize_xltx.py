"""Анонимизатор soliq xltx-выгрузок для коммита в git.

CLI:
    Single-file:  python scripts/anonymize_xltx.py <input.xltx> <output.xltx>
    Batch:        python scripts/anonymize_xltx.py --batch <input_dir> \\
                                                   --output-dir <output_dir> [--force]

Замазывает:
- **ИНН** (9 или 14 подряд цифр) → детерминированный hash той же длины. Один и тот же
  исходный ИНН всегда даёт один и тот же anon-ИНН — формат остаётся валидным,
  cross-file references сохраняются.
- **Имена компаний** (cells содержащие ООО/ОАО/АО/ЗАО/MCHJ/ХК/ХТ/ХУСУСИЙ) →
  ``ОБЕЗЛИЧЕНО NNN`` (счётчик per-file).
- **Суммы** (numeric values > 1000) → ``value / 10``. Сохраняет порядок величин,
  пропорции для KPI-smoke остаются. Реальные суммы в git не утекают.

Формулы (cells с data_type='f') не трогаются — после ``/10`` они пересчитаются
автоматически. Cell с ``data_only=True`` value формулы виделась бы как число —
открываем workbook с ``data_only=False`` чтобы различать formula vs literal.

Batch mode: подхватывает все ``*_full.xltx`` в input_dir, генерирует
``*_anon.xltx`` имена через детерминированную INN-substitution в стеме
(защита от collision'ов когда несколько firms делят префикс). Существующие
anon-файлы пропускаются без ``--force`` — идемпотентный re-run.

Acceptance — anon-файл:
1. Открывается parser'ом без UnsupportedFormatError (sentinel-cells не тронуты —
   они не содержат digits/company markers).
2. ``parse_warnings == []`` (или ⊆ original warnings — anonymization не должна
   плодить новые warning'и).
"""

from __future__ import annotations

import argparse
import hashlib
import logging
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO, format="%(message)s")
_logger = logging.getLogger("anonymize_xltx")

_INN_PATTERN = re.compile(r"(?<!\d)(\d{14}|\d{9})(?!\d)")
_COMPANY_PATTERN = re.compile(
    r"\b(ООО|ОАО|АО|ЗАО|MCHJ|MCJ|АЖ|ХК|ХТ|ХУСУСИЙ)\b",
    re.IGNORECASE,
)
# Порог суммы. Меньше — это, скорее всего, проценты / счётчики / период / число
# страниц, не money. Деление их на 10 ломает header_parser (year=2025 → 202.5).
_AMOUNT_THRESHOLD = 10_000
# Year-like int пропускаем без изменений: 1900 ≤ value ≤ 2100.
_YEAR_MIN, _YEAR_MAX = 1900, 2100
# Длина числа, признаваемая ИНН (numeric, не строка).
_INN_DIGIT_LENGTHS = (9, 14)


def _deterministic_inn(original: str) -> str:
    digest = hashlib.sha256(original.encode("utf-8")).hexdigest()
    digits = "".join(c for c in digest if c.isdigit())
    if len(digits) < len(original):
        digits = (digits + "0" * len(original))[: len(original)]
    return digits[: len(original)]


def _anonymize_string(
    value: str,
    name_counter: list[int],
) -> tuple[str, int, int]:
    inn_replaced = 0

    def _sub_inn(match: re.Match[str]) -> str:
        nonlocal inn_replaced
        inn_replaced += 1
        return _deterministic_inn(match.group(0))

    inn_anon = _INN_PATTERN.sub(_sub_inn, value)

    if _COMPANY_PATTERN.search(inn_anon):
        name_counter[0] += 1
        return f"ОБЕЗЛИЧЕНО {name_counter[0]:03d}", inn_replaced, 1

    return inn_anon, inn_replaced, 0


def _anonymize_number(value: float | int) -> float | int | None:
    """Возвращает заменённое значение или None если cell оставить как есть.

    Контракт:
    - year-like int (1900–2100) → None (header_parser падает на 202.5).
    - numeric с длиной 9 или 14 digits → детерминированный INN-hash (numeric ИНН
      не ловится regex по строкам).
    - abs(value) ≤ AMOUNT_THRESHOLD → None (счётчики / проценты / номера строк).
    - иначе → integer-deлится на 10 (для float аккуратное округление до 2 знаков).
    """
    int_value: int | None = None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        int_value = value
    elif isinstance(value, float) and value.is_integer():
        int_value = int(value)
    if int_value is not None:
        if _YEAR_MIN <= int_value <= _YEAR_MAX:
            return None
        if len(str(abs(int_value))) in _INN_DIGIT_LENGTHS:
            return int(_deterministic_inn(str(abs(int_value))))
    if abs(value) <= _AMOUNT_THRESHOLD:
        return None
    if isinstance(value, int):
        return value // 10
    return round(value / 10, 2)


def _anonymize_filename(input_path: Path) -> str:
    """Преобразует path real ``*_full.xltx`` → anon filename.

    9/14-digit INN в стеме заменяется на детерминированный hash — same
    `_deterministic_inn` что для cell-значений, поэтому cross-file anon-INN
    consistency между filename и содержимым сохраняется.

    Защищает от collision'а когда несколько firms делят префикс
    (``form1_2025_<inn>_full.xltx`` × 4 разных INN → 4 разных anon-имени).
    """
    stem = input_path.stem
    if stem.endswith("_full"):
        stem = stem[: -len("_full")]
    stem = _INN_PATTERN.sub(lambda m: _deterministic_inn(m.group(0)), stem)
    return f"{stem}_anon{input_path.suffix}"


def anonymize_xltx(input_path: Path, output_path: Path) -> dict[str, int]:
    """Анонимизирует xltx-файл, сохраняет в output_path. Возвращает счётчики."""
    wb = load_workbook(filename=str(input_path), data_only=False, read_only=False)
    totals = {"inn": 0, "name": 0, "amount": 0, "cells_scanned": 0}
    name_counter = [0]

    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    totals["cells_scanned"] += 1
                    if cell.value is None:
                        continue
                    if cell.data_type == "f":
                        continue
                    if isinstance(cell.value, str):
                        new_str, inn_n, name_n = _anonymize_string(
                            cell.value, name_counter
                        )
                        if new_str != cell.value:
                            cell.value = new_str
                            totals["inn"] += inn_n
                            totals["name"] += name_n
                    elif isinstance(cell.value, (int, float)) and not isinstance(
                        cell.value, bool
                    ):
                        new_num = _anonymize_number(cell.value)
                        if new_num is not None:
                            cell.value = new_num
                            totals["amount"] += 1
        wb.save(str(output_path))
    finally:
        wb.close()

    return totals


def _run_batch(input_dir: Path, output_dir: Path, force: bool) -> int:
    """Прогоняет анонимизацию по всем ``*_full.xltx`` в input_dir.

    Существующие ``*_anon.xltx`` в output_dir пропускаются без ``--force`` —
    идемпотентный re-run не перетирает уже закоммиченные fixtures (которые
    могли быть сгенерированы более ранней версией скрипта).
    """
    if not input_dir.is_dir():
        _logger.error("[FAIL] --batch %s не директория", input_dir)
        return 2

    output_dir.mkdir(parents=True, exist_ok=True)

    full_files = sorted(input_dir.glob("*_full.xltx"))
    if not full_files:
        _logger.error("[FAIL] в %s нет *_full.xltx", input_dir)
        return 2

    processed = 0
    skipped = 0
    for full_path in full_files:
        anon_name = _anonymize_filename(full_path)
        anon_path = output_dir / anon_name
        if anon_path.exists() and not force:
            _logger.info(
                "[SKIP] %s → %s (уже существует, --force чтобы перезаписать)",
                full_path.name,
                anon_name,
            )
            skipped += 1
            continue
        totals = anonymize_xltx(full_path, anon_path)
        _logger.info(
            "[OK] %s → %s (ИНН=%d, имена=%d, суммы=%d)",
            full_path.name,
            anon_name,
            totals["inn"],
            totals["name"],
            totals["amount"],
        )
        processed += 1

    _logger.info(
        "[DONE] processed=%d skipped=%d total=%d",
        processed,
        skipped,
        len(full_files),
    )
    return 0


def _main() -> int:
    parser = argparse.ArgumentParser(
        description="Анонимизировать soliq xltx для коммита в git",
        epilog=(
            "Single-file: python scripts/anonymize_xltx.py input.xltx output.xltx\n"
            "Batch:       python scripts/anonymize_xltx.py "
            "--batch <input_dir> --output-dir <output_dir> [--force]"
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "input",
        type=Path,
        nargs="?",
        help="входной xltx (single-file mode)",
    )
    parser.add_argument(
        "output",
        type=Path,
        nargs="?",
        help="выходной xltx (single-file mode)",
    )
    parser.add_argument(
        "--batch",
        type=Path,
        default=None,
        help="директория с *_full.xltx для bulk anonymize",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=None,
        help="куда писать *_anon.xltx в batch mode",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="перезаписывать существующие *_anon.xltx (default: skip)",
    )
    args = parser.parse_args()

    if args.batch is not None:
        if args.input is not None or args.output is not None:
            _logger.error("[FAIL] --batch несовместим с positional input/output")
            return 2
        if args.output_dir is None:
            _logger.error("[FAIL] --batch требует --output-dir")
            return 2
        return _run_batch(args.batch, args.output_dir, args.force)

    if args.output_dir is not None:
        _logger.error("[FAIL] --output-dir работает только с --batch")
        return 2

    if args.input is None or args.output is None:
        _logger.error(
            "[FAIL] нужно либо positional input/output, либо --batch + --output-dir"
        )
        return 2

    if not args.input.is_file():
        _logger.error("[FAIL] input не найден: %s", args.input)
        return 2
    if args.output.exists() and args.input.resolve() == args.output.resolve():
        _logger.error("[FAIL] input и output совпадают")
        return 2

    args.output.parent.mkdir(parents=True, exist_ok=True)

    totals = anonymize_xltx(args.input, args.output)
    _logger.info(
        "[OK] anonymized %s → %s (ИНН=%d, имена=%d, суммы=%d, cells=%d)",
        args.input.name,
        args.output.name,
        totals["inn"],
        totals["name"],
        totals["amount"],
        totals["cells_scanned"],
    )
    return 0


if __name__ == "__main__":
    sys.exit(_main())
