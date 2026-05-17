"""Тесты detect_format для всех 5 поддерживаемых форм + UNKNOWN-кейсы."""

from openpyxl import Workbook

from infrastructure.adapters.soliq_xltx.format_detector import (
    SoliqXltxFormat,
    detect_format,
)
from tests.fixtures.soliq_xltx._factories import (
    build_form1_balance_sheet_wb,
    build_form2_income_statement_wb,
    build_profit_tax_wb,
    build_vat_declaration_wb,
    build_vat_registry_wb,
)


def test_detect_vat_declaration() -> None:
    wb = build_vat_declaration_wb()
    assert detect_format(wb) == SoliqXltxFormat.VAT_DECLARATION


def test_detect_vat_declaration_v1_by_narrow_list01() -> None:
    """V1 (10006_41) detection via list01.max_column < 14 (legacy шаблон, 13 колонок).

    Строим minimal workbook вручную — factory build_vat_declaration_wb создаёт
    v2 layout (max_column=16), а openpyxl `cell.value = None` не уменьшает
    max_column. Самый правый cell тут L3 (col 12 — "лист 01" маркер v1).
    """
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "list01"
    # Только sentinel-фраза для VAT detection + один narrow-range cell.
    ws["D8"] = "СВЕДЕНИЯ о плательщике налога на добавленную стоимость"
    ws["L3"] = "лист 01"  # column 12 — типичная позиция в v1
    assert ws.max_column < 14

    assert detect_format(wb) == SoliqXltxFormat.VAT_DECLARATION_V1


def test_detect_vat_registry_ilova() -> None:
    wb = build_vat_registry_wb()
    assert detect_format(wb) == SoliqXltxFormat.VAT_REGISTRY_ILOVA


def test_detect_form2_income_statement() -> None:
    wb = build_form2_income_statement_wb()
    assert detect_format(wb) == SoliqXltxFormat.FORM_2_INCOME_STATEMENT


def test_detect_form1_balance_sheet() -> None:
    wb = build_form1_balance_sheet_wb()
    assert detect_format(wb) == SoliqXltxFormat.FORM_1_BALANCE_SHEET


def test_detect_profit_tax() -> None:
    wb = build_profit_tax_wb()
    assert detect_format(wb) == SoliqXltxFormat.PROFIT_TAX


def test_empty_workbook_is_unknown() -> None:
    wb = Workbook()  # default Sheet, нет list01
    assert detect_format(wb) == SoliqXltxFormat.UNKNOWN


def test_workbook_without_list01_is_unknown() -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    wb.create_sheet("Other")
    assert detect_format(wb) == SoliqXltxFormat.UNKNOWN


def test_workbook_with_list01_but_no_signature_is_unknown() -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "list01"
    ws["B3"] = "Что-то совершенно левое"
    assert detect_format(wb) == SoliqXltxFormat.UNKNOWN


def test_form1_takes_precedence_over_form2_when_both_sentinels_present() -> None:
    # Форма 1 проверяется раньше — если бы B3 содержал оба sentinel'а, выбираем Form 1.
    wb = build_form1_balance_sheet_wb()
    ws = wb["list01"]
    ws["B3"] = "Бухгалтерский баланс - форма № 1 (Отчет о финансовых результатах)"
    assert detect_format(wb) == SoliqXltxFormat.FORM_1_BALANCE_SHEET


# CA-012: разные редакции VAT-декларации Soliq (10006_41 / 10006_45 / 10006_47 …) могут
# смещать sentinel-фразу в другую ячейку или формулировать её иначе. Детектор должен
# опознавать форму по широкому substring-поиску в верхней зоне list01 + structural
# fallback (наличие list02 с числовым G6 и list04 с числовым G7).


def test_detect_vat_declaration_with_sentinel_in_alternate_cell() -> None:
    # Эмуляция другой редакции Soliq, где sentinel-текст лежит не в D8, а, например, в B5.
    wb = build_vat_declaration_wb()
    ws = wb["list01"]
    ws["D8"] = None
    ws["B5"] = "Расчет налога на добавленную стоимость"
    assert detect_format(wb) == SoliqXltxFormat.VAT_DECLARATION


def test_detect_vat_declaration_with_alternate_phrasing() -> None:
    # Другая формулировка ("ДЕКЛАРАЦИЯ ..." вместо "СВЕДЕНИЯ о плательщике ..."),
    # case-insensitive substring-поиск всё равно её ловит.
    wb = build_vat_declaration_wb()
    ws = wb["list01"]
    ws["D8"] = None
    ws["C7"] = "декларация по налогу на добавленную стоимость"
    assert detect_format(wb) == SoliqXltxFormat.VAT_DECLARATION


def test_detect_vat_declaration_via_structural_fallback() -> None:
    # Sentinel вообще не нашёлся, но структура — list02 G6 + list04 G7 числа — совпадает
    # с VAT_DECLARATION → детектим её, а не возвращаем UNKNOWN.
    wb = build_vat_declaration_wb()
    ws = wb["list01"]
    # Полностью затираем заголовочную зону list01 — sentinel-поиск ничего не найдёт.
    for row in ws.iter_rows(min_row=1, max_row=20, max_col=10):
        for cell in row:
            cell.value = None
    assert detect_format(wb) == SoliqXltxFormat.VAT_DECLARATION


def test_unknown_when_no_sentinel_and_no_structure() -> None:
    # Ни sentinel'а, ни числовых list02 G6 / list04 G7 — graceful UNKNOWN, не raise.
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "list01"
    ws["B3"] = "Что-то совершенно левое"
    wb.create_sheet("list02")
    wb.create_sheet("list04")
    assert detect_format(wb) == SoliqXltxFormat.UNKNOWN
