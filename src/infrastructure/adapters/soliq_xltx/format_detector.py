"""Классификация типа .xltx-выгрузки из my3.soliq.uz по содержимому workbook.

Каждая форма (Расчёт НДС, ilova-реестр ЭСФ, Форма №1 баланс, Форма №2 финрезультаты,
Расчёт налога на прибыль) имеет уникальную сигнатуру в первом листе ``list01``.
Детектор не парсит данные — только выбирает специализированный парсер.

Для VAT-деклараций (семейство 10006_*: 10006_41, 10006_45, 10006_47 …) разные
редакции Soliq могут смещать sentinel-фразу в другую ячейку или формулировать её
иначе. Поэтому для VAT_DECLARATION применяется двухуровневая стратегия:

1. Широкий substring-поиск ключевых фраз («налог на добавленную стоимость» +
   маркер документа: «Расчет» / «Расчёт» / «Сведения» / «Декларация») по верхней
   зоне list01 (rows 1–20, cols A–F), case-insensitive.
2. Structural fallback: если в ``list02`` ячейка ``G6`` числовая и в ``list04``
   ячейка ``G7`` числовая (типичная разметка декларации), считаем формат
   VAT_DECLARATION даже без sentinel'а.

Детектор по-прежнему **не смотрит на имя файла** — только содержимое.
"""

from enum import StrEnum

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class SoliqXltxFormat(StrEnum):
    VAT_DECLARATION = "vat_declaration"
    VAT_REGISTRY_ILOVA = "vat_registry_ilova"
    FORM_2_INCOME_STATEMENT = "form_2_income_statement"
    FORM_1_BALANCE_SHEET = "form_1_balance_sheet"
    PROFIT_TAX = "profit_tax"
    UNKNOWN = "unknown"


# Sentinel-фразы (нормализованные substrings), по которым опознаём форму. Сравнение
# case-sensitive — формы Soliq стабильно используют одну и ту же редакцию текста.
_VAT_REGISTRY_SENTINEL = "Реестр счетов-фактур"
_FORM_2_SENTINEL = "Отчет о финансовых результатах"
_FORM_1_SENTINEL = "Бухгалтерский баланс"
_PROFIT_TAX_SENTINEL = "налога на прибыль"

# VAT-декларация ищется широко: фраза про НДС + любой из document-маркеров.
# Substring без падежа — реальные шаблоны Soliq используют «налога», «налогом», «налоге».
_VAT_PHRASE = "на добавленную стоимость"
_VAT_DOC_MARKERS = ("расчет", "расчёт", "сведения", "декларация")
# Зона поиска sentinel'а в list01 (rows × cols).
_VAT_SCAN_ROWS = 20
_VAT_SCAN_COLS = 6


def detect_format(wb: Workbook) -> SoliqXltxFormat:
    """Определить тип выгрузки по сигнатурным ячейкам list01.

    Возвращает ``UNKNOWN``, если list01 отсутствует или ни одна сигнатура не совпала.
    """
    if "list01" not in wb.sheetnames:
        return SoliqXltxFormat.UNKNOWN
    ws = wb["list01"]

    def text(coord: str) -> str:
        v = ws[coord].value
        return str(v) if v is not None else ""

    # Ilova-реестр: либо B9 (Таблица №1, закупки), либо list02 B9 (Таблица №2, продажи).
    # Для устойчивости проверяем оба возможных места (file может быть только Таблицей №2).
    # Проверяем раньше VAT_DECLARATION — иначе widened VAT-поиск может ложно сматчить
    # «Расчёт налога на добавленную стоимость» в подзаголовке ilova.
    if _VAT_REGISTRY_SENTINEL in text("B9") or _VAT_REGISTRY_SENTINEL in text("B8"):
        return SoliqXltxFormat.VAT_REGISTRY_ILOVA

    if _FORM_1_SENTINEL in text("B3"):
        return SoliqXltxFormat.FORM_1_BALANCE_SHEET

    if _FORM_2_SENTINEL in text("B3"):
        return SoliqXltxFormat.FORM_2_INCOME_STATEMENT

    if _PROFIT_TAX_SENTINEL in text("B2"):
        return SoliqXltxFormat.PROFIT_TAX

    if _is_vat_declaration(wb, ws):
        return SoliqXltxFormat.VAT_DECLARATION

    return SoliqXltxFormat.UNKNOWN


def _is_vat_declaration(wb: Workbook, list01: Worksheet) -> bool:
    """Двухуровневая детекция VAT-декларации: sentinel в list01 ИЛИ structural fallback."""
    if _scan_vat_sentinel(list01):
        return True
    return _has_vat_declaration_structure(wb)


def _scan_vat_sentinel(ws: Worksheet) -> bool:
    """Ищет фразу про НДС + document-маркер в верхней зоне list01.

    Случай-нечувствительно склеиваем все непустые cells в зоне scan и ищем по тексту.
    Это устойчиво к разным редакциям формы (10006_41/45/47 …) и к смещению sentinel.
    """
    parts: list[str] = []
    for row in ws.iter_rows(
        min_row=1, max_row=_VAT_SCAN_ROWS, min_col=1, max_col=_VAT_SCAN_COLS
    ):
        for cell in row:
            v = cell.value
            if v is None:
                continue
            parts.append(str(v))
    if not parts:
        return False
    blob = " ".join(parts).lower()
    if _VAT_PHRASE not in blob:
        return False
    return any(marker in blob for marker in _VAT_DOC_MARKERS)


def _has_vat_declaration_structure(wb: Workbook) -> bool:
    """Structural fallback: list02 G6 и list04 G7 — числовые (типичная разметка)."""
    if "list02" not in wb.sheetnames or "list04" not in wb.sheetnames:
        return False
    list02 = wb["list02"]
    list04 = wb["list04"]
    return _is_number(list02["G6"].value) and _is_number(list04["G7"].value)


def _is_number(value: object) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)
