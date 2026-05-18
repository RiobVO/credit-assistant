"""Тесты parse_header — извлечение шапки из xltx-форм.

После best-effort рефактора (CA-014 hardening): row/cell-level ошибки
становятся parse_warnings + None, не raise. Raise остаются только для
структурных ошибок (отсутствие list01, неподдерживаемый формат)."""

from datetime import date
from decimal import Decimal

import pytest

from domain.value_objects.inn import INN
from infrastructure.adapters.soliq_xltx.errors import UnsupportedFormatError
from infrastructure.adapters.soliq_xltx.format_detector import SoliqXltxFormat
from infrastructure.adapters.soliq_xltx.header_parser import (
    parse_header,
    parse_unit_multiplier,
)
from tests.fixtures.soliq_xltx._factories import (
    build_form1_balance_sheet_wb,
    build_form2_income_statement_wb,
    build_profit_tax_wb,
    build_vat_declaration_wb,
)


class TestVatDeclarationHeader:
    def test_full_header_parsed(self) -> None:
        wb = build_vat_declaration_wb(
            inn=306399449,
            organization_name='"AZ RUHDIL SAVDO" MCHJ',
            period_year=2026,
            period_kind="месяц",
            submitted_at="20.04.2026",
        )
        h = parse_header(wb, SoliqXltxFormat.VAT_DECLARATION)
        assert h.borrower_inn == INN("306399449")
        assert h.organization_name == '"AZ RUHDIL SAVDO" MCHJ'
        assert h.period_year == 2026
        assert h.period_kind == "month"
        assert h.period_index is None  # для VAT declaration номер не в шапке
        assert h.submitted_at == date(2026, 4, 20)
        assert h.parse_warnings == []


class TestVatDeclarationHeaderV1:
    """T0.5 Bug C: legacy 10006_41 layout — координаты сдвинуты влево на 2-4 колонки."""

    def test_v1_header_parsed_from_legacy_coords(self) -> None:
        # Строим v1 layout вручную — sentinel в D8, координаты по v1 mapping:
        # INN=C3, name=F9, year=K5, period_kind=I5, date=F17.
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "list01"
        ws["B3"] = "ИНН"
        ws["C3"] = 305002665
        ws["I5"] = "месяц"
        ws["K5"] = 2025
        ws["D8"] = "СВЕДЕНИЯ о плательщике налога на добавленную стоимость"
        ws["F9"] = '"ZAMIN NOZ NEMATLARI" MAS\'ULIYATI CHEKLANGAN JAMIYAT'
        ws["F17"] = "20.01.2026"
        # max_column не должен превысить v1 threshold (14): самый правый — K (col 11).
        assert ws.max_column < 14

        h = parse_header(wb, SoliqXltxFormat.VAT_DECLARATION_V1)
        assert h.borrower_inn == INN("305002665")
        assert h.organization_name == '"ZAMIN NOZ NEMATLARI" MAS\'ULIYATI CHEKLANGAN JAMIYAT'
        assert h.period_year == 2025
        assert h.period_kind == "month"
        assert h.submitted_at == date(2026, 1, 20)
        assert h.parse_warnings == []

    def test_inn_as_string_with_whitespace(self) -> None:
        wb = build_vat_declaration_wb(inn="  306399449  ")
        h = parse_header(wb, SoliqXltxFormat.VAT_DECLARATION)
        assert h.borrower_inn == INN("306399449")

    def test_quarter_kind(self) -> None:
        wb = build_vat_declaration_wb(period_kind="квартал")
        h = parse_header(wb, SoliqXltxFormat.VAT_DECLARATION)
        assert h.period_kind == "quarter"

    def test_invalid_inn_warns_and_returns_none(self) -> None:
        wb = build_vat_declaration_wb(inn=12345)  # короче 9 цифр
        h = parse_header(wb, SoliqXltxFormat.VAT_DECLARATION)
        assert h.borrower_inn is None
        assert any("INN" in w and "D3" in w for w in h.parse_warnings)
        # Остальные поля парсятся как обычно
        assert h.period_year == 2026

    def test_missing_org_name_warns_and_returns_none(self) -> None:
        wb = build_vat_declaration_wb()
        wb["list01"]["H9"].value = None
        h = parse_header(wb, SoliqXltxFormat.VAT_DECLARATION)
        assert h.organization_name is None
        assert any("H9" in w for w in h.parse_warnings)

    def test_missing_year_warns_and_returns_none(self) -> None:
        wb = build_vat_declaration_wb()
        wb["list01"]["O5"].value = None
        h = parse_header(wb, SoliqXltxFormat.VAT_DECLARATION)
        assert h.period_year is None
        assert any("O5" in w for w in h.parse_warnings)

    def test_missing_submitted_at_is_silent_none(self) -> None:
        # submitted_at и так был optional — не warning, просто None.
        wb = build_vat_declaration_wb()
        wb["list01"]["H17"].value = None
        h = parse_header(wb, SoliqXltxFormat.VAT_DECLARATION)
        assert h.submitted_at is None
        assert h.parse_warnings == []

    def test_multiple_invalid_fields_collect_all_warnings(self) -> None:
        wb = build_vat_declaration_wb(inn=12345)
        wb["list01"]["H9"].value = None
        wb["list01"]["O5"].value = None
        h = parse_header(wb, SoliqXltxFormat.VAT_DECLARATION)
        assert h.borrower_inn is None
        assert h.organization_name is None
        assert h.period_year is None
        assert len(h.parse_warnings) >= 3

    def test_empty_list01_returns_all_none_with_warnings(self) -> None:
        wb = build_vat_declaration_wb()
        for row in wb["list01"].iter_rows(min_row=1, max_row=20, max_col=20):
            for cell in row:
                cell.value = None
        h = parse_header(wb, SoliqXltxFormat.VAT_DECLARATION)
        assert h.borrower_inn is None
        assert h.organization_name is None
        assert h.period_year is None
        assert h.parse_warnings  # не пустой


class TestForm2Header:
    def test_full_header(self) -> None:
        wb = build_form2_income_statement_wb(inn=306399449, period_year=2025, period_quarter=4)
        h = parse_header(wb, SoliqXltxFormat.FORM_2_INCOME_STATEMENT)
        assert h.borrower_inn == INN("306399449")
        assert h.period_year == 2025
        assert h.period_kind == "quarter"
        assert h.period_index == 4
        assert h.parse_warnings == []


class TestForm1Header:
    def test_full_header(self) -> None:
        wb = build_form1_balance_sheet_wb(inn=306399449, period_year=2025, period_quarter=4)
        h = parse_header(wb, SoliqXltxFormat.FORM_1_BALANCE_SHEET)
        assert h.borrower_inn == INN("306399449")
        assert h.period_year == 2025
        assert h.period_index == 4


class TestProfitTaxHeader:
    def test_full_header(self) -> None:
        wb = build_profit_tax_wb(inn=306399449, period_year=2026, period_quarter=1)
        h = parse_header(wb, SoliqXltxFormat.PROFIT_TAX)
        assert h.borrower_inn == INN("306399449")
        assert h.period_year == 2026
        assert h.period_index == 1


class TestUnsupported:
    """Структурные ошибки остаются raise — best-effort только на cell-уровне."""

    def test_registry_format_not_supported(self) -> None:
        wb = build_vat_declaration_wb()  # любой workbook
        with pytest.raises(UnsupportedFormatError):
            parse_header(wb, SoliqXltxFormat.VAT_REGISTRY_ILOVA)

    def test_unknown_format_not_supported(self) -> None:
        wb = build_vat_declaration_wb()
        with pytest.raises(UnsupportedFormatError):
            parse_header(wb, SoliqXltxFormat.UNKNOWN)


class TestUnitMultiplier:
    """T2.1 CA-028: parse_unit_multiplier — динамическая детекция «тыс. сум.» /
    «млн сум.» / «сум.» (полные) из B23 (FORM_1) / B24 (FORM_2).

    Возвращает (Decimal multiplier, str | None warning). Unknown text / empty
    cell — warning + fallback ×1000 (банк-friendly: не теряем файл).

    PROFIT_TAX / VAT_DECLARATION / VAT_REGISTRY_ILOVA — UnsupportedFormatError
    (структурная защита: их парсеры используют ×1, helper не нужен)."""

    def test_form2_default_thousand(self) -> None:
        """Default «Единица измерения, тыс. сум.» → ×1000, без warning."""
        wb = build_form2_income_statement_wb()
        multiplier, warning = parse_unit_multiplier(
            wb, SoliqXltxFormat.FORM_2_INCOME_STATEMENT
        )
        assert multiplier == Decimal(1000)
        assert warning is None

    def test_form2_million(self) -> None:
        """«млн сум.» → ×1_000_000."""
        wb = build_form2_income_statement_wb(unit_text="Единица измерения, млн. сум.")
        multiplier, warning = parse_unit_multiplier(
            wb, SoliqXltxFormat.FORM_2_INCOME_STATEMENT
        )
        assert multiplier == Decimal(1_000_000)
        assert warning is None

    def test_form2_full_sums(self) -> None:
        """«сум.» (без тыс./млн) → ×1 (полные сум)."""
        wb = build_form2_income_statement_wb(unit_text="Единица измерения, сум.")
        multiplier, warning = parse_unit_multiplier(
            wb, SoliqXltxFormat.FORM_2_INCOME_STATEMENT
        )
        assert multiplier == Decimal(1)
        assert warning is None

    def test_form2_unknown_text_fallback_with_warning(self) -> None:
        """Не распознан → fallback ×1000 + warning (банк-friendly)."""
        wb = build_form2_income_statement_wb(unit_text="мусор без распознавания")
        multiplier, warning = parse_unit_multiplier(
            wb, SoliqXltxFormat.FORM_2_INCOME_STATEMENT
        )
        assert multiplier == Decimal(1000)
        assert warning is not None
        assert "B24" in warning

    def test_form2_empty_cell_fallback_with_warning(self) -> None:
        """B24 пустой → fallback ×1000 + warning."""
        wb = build_form2_income_statement_wb(unit_text=None)
        multiplier, warning = parse_unit_multiplier(
            wb, SoliqXltxFormat.FORM_2_INCOME_STATEMENT
        )
        assert multiplier == Decimal(1000)
        assert warning is not None
        assert "B24" in warning

    def test_form1_default_thousand_from_b23(self) -> None:
        """FORM_1 B23 (отличается от FORM_2 B24) → ×1000."""
        wb = build_form1_balance_sheet_wb()
        multiplier, warning = parse_unit_multiplier(
            wb, SoliqXltxFormat.FORM_1_BALANCE_SHEET
        )
        assert multiplier == Decimal(1000)
        assert warning is None

    def test_form1_million_from_b23(self) -> None:
        """FORM_1 «млн сум.» в B23 → ×1_000_000."""
        wb = build_form1_balance_sheet_wb(unit_text="Единица измерения, млн. сум.")
        multiplier, warning = parse_unit_multiplier(
            wb, SoliqXltxFormat.FORM_1_BALANCE_SHEET
        )
        assert multiplier == Decimal(1_000_000)
        assert warning is None

    def test_case_insensitive_matching(self) -> None:
        """Регистр не важен — Soliq может использовать «ТЫС.» / «Млн»."""
        wb = build_form2_income_statement_wb(unit_text="ЕДИНИЦА ИЗМЕРЕНИЯ, ТЫС. СУМ.")
        multiplier, _ = parse_unit_multiplier(
            wb, SoliqXltxFormat.FORM_2_INCOME_STATEMENT
        )
        assert multiplier == Decimal(1000)

    def test_profit_tax_unsupported(self) -> None:
        """PROFIT_TAX → UnsupportedFormatError (его парсер ×1, helper не для него)."""
        wb = build_profit_tax_wb()
        with pytest.raises(UnsupportedFormatError):
            parse_unit_multiplier(wb, SoliqXltxFormat.PROFIT_TAX)

    def test_vat_declaration_unsupported(self) -> None:
        """VAT_DECLARATION → UnsupportedFormatError (×1 уже)."""
        wb = build_vat_declaration_wb()
        with pytest.raises(UnsupportedFormatError):
            parse_unit_multiplier(wb, SoliqXltxFormat.VAT_DECLARATION)
