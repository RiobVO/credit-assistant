"""Извлечение шапки из xltx-форм my3.soliq.uz: ИНН, организация, период.

Координаты шапки разные для каждой формы — диспатч по ``SoliqXltxFormat``.
Ilova-реестр (Приложение №4) собственной шапки не имеет, parse_header его не
обслуживает — она прилагается к декларации, к которой и относится файл.

Best-effort семантика: cell-level ошибки попадают в ``parse_warnings`` и
возвращают ``None`` для соответствующего поля; raise остаётся только для
структурных ошибок (отсутствие list01, не та форма) — они блокируют сам
вызов parse_header. Это закрывает CA-014 на уровне шапки.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Literal

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from domain.value_objects.inn import INN, InvalidInnError
from infrastructure.adapters.soliq_xltx.errors import UnsupportedFormatError
from infrastructure.adapters.soliq_xltx.format_detector import SoliqXltxFormat

PeriodKind = Literal["month", "quarter", "annual"]

_logger = logging.getLogger(__name__)


@dataclass(slots=True)
class SoliqXltxHeader:
    """Унифицированная шапка xltx-формы (best-effort).

    Все поля nullable: cell-level ошибка → ``None`` + запись в
    ``parse_warnings``. Downstream-код решает, что делать с пропусками.

    ``period_index`` — номер месяца (1-12) либо квартала (1-4); ``None``, если
    форма годовая или конкретный номер не указан в шапке.
    """

    borrower_inn: INN | None
    organization_name: str | None
    period_year: int | None
    period_kind: PeriodKind | None
    period_index: int | None
    submitted_at: date | None
    parse_warnings: list[str] = field(default_factory=list)


def parse_header(wb: Workbook, fmt: SoliqXltxFormat) -> SoliqXltxHeader:
    """Прочитать шапку формы по правилам ``fmt``.

    Raise остаётся только для структурных ошибок: нет list01, не та форма.
    """
    if "list01" not in wb.sheetnames:
        raise UnsupportedFormatError(wb.sheetnames, "no list01")
    ws = wb["list01"]

    if fmt is SoliqXltxFormat.VAT_DECLARATION:
        return _parse_vat_declaration_header(ws)
    if fmt is SoliqXltxFormat.FORM_2_INCOME_STATEMENT:
        return _parse_form2_header(ws)
    if fmt is SoliqXltxFormat.FORM_1_BALANCE_SHEET:
        return _parse_form1_header(ws)
    if fmt is SoliqXltxFormat.PROFIT_TAX:
        return _parse_profit_tax_header(ws)
    raise UnsupportedFormatError(wb.sheetnames, f"parse_header не определён для {fmt}")


def _parse_vat_declaration_header(ws: Worksheet) -> SoliqXltxHeader:
    warnings: list[str] = []
    inn = _read_inn(ws, "D3", warnings)
    org = _read_text(ws, "H9", warnings, label="organization name")
    year = _read_int(ws, "O5", warnings, label="period_year")
    kind = _read_period_kind(ws, "K5", warnings)
    submitted = _read_date(ws, "H17")
    return SoliqXltxHeader(
        borrower_inn=inn,
        organization_name=org,
        period_year=year,
        period_kind=kind,
        period_index=None,
        submitted_at=submitted,
        parse_warnings=warnings,
    )


def _parse_form2_header(ws: Worksheet) -> SoliqXltxHeader:
    warnings: list[str] = []
    inn = _read_inn(ws, "I18", warnings)
    org = _read_text(ws, "C8", warnings, label="organization name")
    year = _read_int(ws, "C5", warnings, label="period_year")
    quarter = _read_optional_int(ws, "E5")
    return SoliqXltxHeader(
        borrower_inn=inn,
        organization_name=org,
        period_year=year,
        period_kind="quarter" if quarter is not None else "annual",
        period_index=quarter,
        submitted_at=None,
        parse_warnings=warnings,
    )


def _parse_form1_header(ws: Worksheet) -> SoliqXltxHeader:
    warnings: list[str] = []
    inn = _read_inn(ws, "I17", warnings)
    org = _read_text(ws, "C7", warnings, label="organization name")
    year = _read_int(ws, "C4", warnings, label="period_year")
    quarter = _read_optional_int(ws, "E4")
    return SoliqXltxHeader(
        borrower_inn=inn,
        organization_name=org,
        period_year=year,
        period_kind="quarter" if quarter is not None else "annual",
        period_index=quarter,
        submitted_at=None,
        parse_warnings=warnings,
    )


def _parse_profit_tax_header(ws: Worksheet) -> SoliqXltxHeader:
    warnings: list[str] = []
    inn = _read_inn(ws, "C4", warnings)
    org = _read_text(ws, "G8", warnings, label="organization name")
    year = _read_int(ws, "K6", warnings, label="period_year")
    quarter = _read_optional_int(ws, "G6")
    return SoliqXltxHeader(
        borrower_inn=inn,
        organization_name=org,
        period_year=year,
        period_kind="quarter" if quarter is not None else "annual",
        period_index=quarter,
        submitted_at=None,
        parse_warnings=warnings,
    )


def _warn(warnings: list[str], sheet: str, coord: str, reason: str) -> None:
    """Добавить запись в parse_warnings + structured лог."""
    msg = f"{sheet}!{coord}: {reason}"
    warnings.append(msg)
    _logger.warning("xltx.header_field_skipped sheet=%s coord=%s reason=%s", sheet, coord, reason)


def _read_inn(ws: Worksheet, coord: str, warnings: list[str]) -> INN | None:
    raw = ws[coord].value
    if raw is None:
        _warn(warnings, ws.title, coord, "INN cell is empty")
        return None
    if isinstance(raw, float):
        if not raw.is_integer():
            _warn(warnings, ws.title, coord, f"INN must be integer, got {raw}")
            return None
        text = str(int(raw))
    else:
        text = str(raw).strip()
    try:
        return INN(text)
    except InvalidInnError as exc:
        _warn(warnings, ws.title, coord, f"invalid INN: {exc}")
        return None


def _read_text(ws: Worksheet, coord: str, warnings: list[str], *, label: str) -> str | None:
    raw = ws[coord].value
    if raw is None:
        _warn(warnings, ws.title, coord, f"{label} missing")
        return None
    text = str(raw).strip()
    if not text:
        _warn(warnings, ws.title, coord, f"{label} missing (empty)")
        return None
    return text


def _read_int(ws: Worksheet, coord: str, warnings: list[str], *, label: str) -> int | None:
    raw = ws[coord].value
    if raw is None:
        _warn(warnings, ws.title, coord, f"{label} missing")
        return None
    if isinstance(raw, int) and not isinstance(raw, bool):
        return raw
    if isinstance(raw, float) and raw.is_integer():
        return int(raw)
    if isinstance(raw, str):
        try:
            return int(raw.strip())
        except ValueError:
            _warn(warnings, ws.title, coord, f"{label}: not an integer ({raw!r})")
            return None
    _warn(warnings, ws.title, coord, f"{label}: unexpected type {type(raw).__name__}")
    return None


def _read_optional_int(ws: Worksheet, coord: str) -> int | None:
    """Тихий optional int (не пишет warning — поле само по себе optional)."""
    raw = ws[coord].value
    if raw is None:
        return None
    if isinstance(raw, int) and not isinstance(raw, bool):
        return raw
    if isinstance(raw, float) and raw.is_integer():
        return int(raw)
    if isinstance(raw, str) and raw.strip().isdigit():
        return int(raw.strip())
    return None


def _read_period_kind(ws: Worksheet, coord: str, warnings: list[str]) -> PeriodKind | None:
    raw = ws[coord].value
    if raw is None:
        return "annual"
    text = str(raw).strip().lower()
    if "месяц" in text:
        return "month"
    if "кварт" in text:
        return "quarter"
    if "год" in text:
        return "annual"
    _warn(warnings, ws.title, coord, f"unrecognized period kind: {raw!r}")
    return None


def _read_date(ws: Worksheet, coord: str) -> date | None:
    """Тихий optional date (поле само по себе optional)."""
    raw = ws[coord].value
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    if isinstance(raw, str):
        try:
            return datetime.strptime(raw.strip(), "%d.%m.%Y").date()
        except ValueError:
            return None
    return None
