"""Парсер ilova-приложения №4 к Расчёту НДС: реестр счетов-фактур построчно.

В отличие от CSV из e-factura.uz (parser в ``esf_csv``), этот файл содержит
**сумму НДС по каждой ЭСФ** — что и закрывает дилемму с ``esf_seller_vat_total``,
описанную в ADR 0004.

Best-effort семантика (CA-014 hardening): cell-level ошибки → warning + None
по полю; если критичные поля (name, date, amount, vat) непригодны — строка
скипается с записью в ``parse_warnings``. Парсер никогда не raises на
данных — только на структурных ошибках (не та форма, отсутствие листов).

Структура (по реальной выгрузке папы, март 2026, 503 продажи + 53 закупки):
- list01 (Таблица №1, **закупки**): шапка R11, данные с R15.
- list02 (Таблица №2, **продажи**): шапка R11, данные с R15.

Колонки data rows:
- B: №
- C: Наименование контрагента (required)
- D: ИНН контрагента (опционально — пусто для розничных покупателей-ПИНФЛ)
- E: Номер счёта-фактуры
- F: Дата счёта-фактуры (формат DD.MM.YYYY либо datetime)
- G: Стоимость поставки без НДС
- H: Сумма НДС
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from domain.value_objects.money import Currency, Money
from infrastructure.adapters.soliq_xltx.errors import UnsupportedFormatError
from infrastructure.adapters.soliq_xltx.format_detector import (
    SoliqXltxFormat,
    detect_format,
)

_logger = logging.getLogger(__name__)


@dataclass(frozen=True, slots=True)
class VatRegistryRow:
    seq_no: int
    counterparty_name: str
    counterparty_inn: str | None
    invoice_no: str
    invoice_date: date
    amount_excl_vat: Money
    vat_amount: Money


@dataclass(slots=True)
class VatRegistryData:
    """Декомпозиция реестра ЭСФ на закупки и продажи.

    ``sales_vat_total`` — сумма НДС по всем строкам ``sales`` (агрегат за период).
    Это и есть ``esf_seller_vat_total`` для домена в Day 2.

    ``skipped_rows_count`` — счётчик строк, которые best-effort-парсер пропустил
    из-за непригодных critical-cells (name, date, amount, vat). Каждая такая
    строка добавляет запись в ``parse_warnings`` с (sheet, row, reason) для
    показа в UI и идёт в logging.WARN. Структурные ошибки (не та форма,
    отсутствие листов) по-прежнему raise — глотаем только row-level.
    """

    sales: list[VatRegistryRow] = field(default_factory=list)
    purchases: list[VatRegistryRow] = field(default_factory=list)
    sales_vat_total: Money = field(default_factory=lambda: Money(Decimal(0), Currency.UZS))
    purchases_vat_total: Money = field(default_factory=lambda: Money(Decimal(0), Currency.UZS))
    sales_amount_total: Money = field(default_factory=lambda: Money(Decimal(0), Currency.UZS))
    purchases_amount_total: Money = field(default_factory=lambda: Money(Decimal(0), Currency.UZS))
    skipped_rows_count: int = 0
    parse_warnings: list[str] = field(default_factory=list)


_DATA_START_ROW = 15

# Маркер розничной операции через онлайн-кассу (Onlayn NKM, Onlayn-NKM, NKM, …).
# В таких rows counterparty name (C) пустой — покупатель анонимный розничный,
# вместо name в counterparty заносится placeholder. См. T0.5.
_NKM_MARKER = "NKM"
_NKM_COUNTERPARTY_PLACEHOLDER = "[NKM Retail]"


def parse_vat_registry(wb: Workbook) -> VatRegistryData:
    """Прочитать ilova-реестр (list01 + list02) с детализацией ЭСФ.

    Best-effort: row-level ошибки логируются и пропускаются с записью в
    ``parse_warnings``; структурные (не тот формат) — пробрасываются.
    """
    fmt = detect_format(wb)
    if fmt is not SoliqXltxFormat.VAT_REGISTRY_ILOVA:
        raise UnsupportedFormatError(wb.sheetnames, f"expected VAT_REGISTRY_ILOVA, got {fmt}")

    warnings: list[str] = []
    purchases, purchases_skipped = (
        _read_rows(wb["list01"], warnings) if "list01" in wb.sheetnames else ([], 0)
    )
    sales, sales_skipped = (
        _read_rows(wb["list02"], warnings) if "list02" in wb.sheetnames else ([], 0)
    )

    return VatRegistryData(
        sales=sales,
        purchases=purchases,
        sales_vat_total=_sum_money([r.vat_amount for r in sales]),
        purchases_vat_total=_sum_money([r.vat_amount for r in purchases]),
        sales_amount_total=_sum_money([r.amount_excl_vat for r in sales]),
        purchases_amount_total=_sum_money([r.amount_excl_vat for r in purchases]),
        skipped_rows_count=sales_skipped + purchases_skipped,
        parse_warnings=warnings,
    )


def _read_rows(ws: Worksheet, warnings: list[str]) -> tuple[list[VatRegistryRow], int]:
    """Прочитать data rows реестра начиная с _DATA_START_ROW.

    Останавливается на первой пустой строке (B-cell None и C-cell None) — Soliq
    кладёт пустые строки-заполнители в конце листа.

    Best-effort: если row не парсится, _parse_row возвращает None и причину;
    добавляем в warnings + logging.WARN, идём дальше.
    """
    rows: list[VatRegistryRow] = []
    skipped = 0
    max_row = ws.max_row
    for r in range(_DATA_START_ROW, max_row + 1):
        seq_raw = ws.cell(row=r, column=2).value
        name_raw = ws.cell(row=r, column=3).value
        # Stop conditions:
        # (a) classic: seq и name пустые → конец реестра.
        # (b) trailing partial: seq есть, но **все** payload-cells (name + amount +
        #     vat) пусты → это «огрызки» в конце листа (Soliq оставляет
        #     зарезервированные номера строк без данных). См. T0.5 Bug B.
        if seq_raw is None and name_raw is None:
            break
        if name_raw is None:
            amount_raw = ws.cell(row=r, column=7).value
            vat_raw = ws.cell(row=r, column=8).value
            invoice_raw = ws.cell(row=r, column=5).value
            if amount_raw is None and vat_raw is None and invoice_raw is None:
                break
        try:
            row, skip_reason = _parse_row(ws, r)
        except Exception as exc:
            # Защита от непредсказуемых ошибок openpyxl/Decimal — не должны падать.
            row, skip_reason = None, f"unexpected error: {exc!r}"
        if row is None:
            skipped += 1
            msg = f"{ws.title}!row {r}: {skip_reason or 'invalid row'}"
            warnings.append(msg)
            _logger.warning(
                "xltx.registry_row_skipped sheet=%s row=%d reason=%s",
                ws.title,
                r,
                skip_reason,
            )
        else:
            rows.append(row)
    return rows, skipped


def _parse_row(ws: Worksheet, r: int) -> tuple[VatRegistryRow | None, str | None]:
    """Прочитать одну row реестра.

    Возвращает ``(row, None)`` при успехе или ``(None, reason)`` если хотя бы
    одно critical-поле (name / invoice_no / date / amount / vat) не разобралось.
    seq_no при поломке заменяется на 0 — это не critical для downstream.

    NKM (онлайн-касса, T0.5 Bug A): если invoice_no (E) содержит маркер ``NKM``,
    counterparty_name может быть пустым — это розничная анонимная продажа через
    онлайн-кассу, name заменяется на ``[NKM Retail]`` placeholder.
    """
    seq_no = _int_cell(ws, r, 2) or 0
    name = _str_cell(ws, r, 3)
    inn = _str_cell(ws, r, 4)  # optional
    invoice_no = _str_cell(ws, r, 5)
    invoice_date = _date_cell(ws, r, 6)
    amount_excl_vat = _money_cell(ws, r, 7)
    vat_amount = _money_cell(ws, r, 8)

    is_nkm = invoice_no is not None and _NKM_MARKER in invoice_no.upper()

    if name is None:
        if is_nkm:
            name = _NKM_COUNTERPARTY_PLACEHOLDER
        else:
            return None, f"{ws.cell(row=r, column=3).coordinate}: name is empty"
    if invoice_no is None:
        return None, f"{ws.cell(row=r, column=5).coordinate}: invoice_no is empty"
    if invoice_date is None:
        raw = ws.cell(row=r, column=6).value
        return None, f"{ws.cell(row=r, column=6).coordinate}: invalid date {raw!r}"
    if amount_excl_vat is None or vat_amount is None:
        return None, "amount/vat cell is invalid"

    return (
        VatRegistryRow(
            seq_no=seq_no,
            counterparty_name=name,
            counterparty_inn=inn,
            invoice_no=invoice_no,
            invoice_date=invoice_date,
            amount_excl_vat=amount_excl_vat,
            vat_amount=vat_amount,
        ),
        None,
    )


def _int_cell(ws: Worksheet, r: int, c: int) -> int | None:
    raw = ws.cell(row=r, column=c).value
    if isinstance(raw, bool):  # bool — подкласс int, не нужен здесь
        return None
    if isinstance(raw, int):
        return raw
    if isinstance(raw, float) and raw.is_integer():
        return int(raw)
    return None


def _str_cell(ws: Worksheet, r: int, c: int) -> str | None:
    raw = ws.cell(row=r, column=c).value
    if raw is None:
        return None
    text = str(raw).strip()
    return text or None


def _date_cell(ws: Worksheet, r: int, c: int) -> date | None:
    raw = ws.cell(row=r, column=c).value
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    if isinstance(raw, str):
        try:
            return datetime.strptime(raw.strip(), "%d.%m.%Y").date()
        except ValueError:
            return None
    return None


def _money_cell(ws: Worksheet, r: int, c: int) -> Money | None:
    raw = ws.cell(row=r, column=c).value
    if raw is None:
        return Money(Decimal(0), Currency.UZS)
    if isinstance(raw, str):
        text = raw.strip()
        if not text:
            return Money(Decimal(0), Currency.UZS)
        try:
            return Money(Decimal(text.replace(",", ".")), Currency.UZS)
        except Exception:
            return None
    if isinstance(raw, bool):
        return None
    if isinstance(raw, (int, float)):
        return Money(Decimal(str(raw)), Currency.UZS)
    return None


def _sum_money(values: list[Money]) -> Money:
    if not values:
        return Money(Decimal(0), Currency.UZS)
    total = Money(Decimal(0), Currency.UZS)
    for v in values:
        total = total + v
    return total
