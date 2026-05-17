"""Facade-адаптер: открыть .xltx, классифицировать формат, диспетч на парсер.

Поддержанные парсеры:
- ``parse_vat_declaration`` для Расчёта НДС (8 листов)
- ``parse_vat_registry`` для ilova-приложения №4 (10 листов)
- ``parse_form2`` для Формы №2 «Отчёт о финансовых результатах» (3 листа)
- ``parse_form1`` для Формы №1 «Бухгалтерский баланс» (4 листа)

Profit Tax распознаётся ``detect_format`` и читает шапку через
``parse_header``, но специализированного парсера для него ещё нет
(TODO[CA-029] — отдельным заходом после FORM_1).
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import BinaryIO

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from infrastructure.adapters.soliq_xltx.errors import UnsupportedFormatError
from infrastructure.adapters.soliq_xltx.form1_parser import (
    Form1BalanceSheetData,
    parse_form1,
)
from infrastructure.adapters.soliq_xltx.form2_parser import (
    Form2IncomeStatementData,
    parse_form2,
)
from infrastructure.adapters.soliq_xltx.format_detector import (
    SoliqXltxFormat,
    detect_format,
)
from infrastructure.adapters.soliq_xltx.vat_declaration_parser import (
    VatDeclarationData,
    parse_vat_declaration,
)
from infrastructure.adapters.soliq_xltx.vat_registry_parser import (
    VatRegistryData,
    parse_vat_registry,
)

ParsedSoliqXltx = (
    VatDeclarationData
    | VatRegistryData
    | Form2IncomeStatementData
    | Form1BalanceSheetData
)


class SoliqXltxAdapter:
    """Точка входа для парсинга .xltx-выгрузок my3.soliq.uz.

    ``parse(source)`` принимает либо ``Path``/строку с файловым путём, либо
    bytes/BinaryIO; внутри использует ``openpyxl.load_workbook(data_only=True)``
    — берём cell values, а не формулы (в ilova ИТОГО-ячейки содержат =SUM(...)).
    """

    def parse(self, source: str | Path | bytes | BinaryIO) -> ParsedSoliqXltx:
        wb = self._load(source)
        try:
            return self._dispatch(wb)
        finally:
            wb.close()

    def detect(self, source: str | Path | bytes | BinaryIO) -> SoliqXltxFormat:
        """Определить формат без полного парсинга — полезно перед routing-ом."""
        wb = self._load(source)
        try:
            return detect_format(wb)
        finally:
            wb.close()

    @staticmethod
    def _load(source: str | Path | bytes | BinaryIO) -> Workbook:
        if isinstance(source, (str, Path)):
            return load_workbook(filename=str(source), data_only=True, read_only=False)
        if isinstance(source, bytes):
            return load_workbook(filename=BytesIO(source), data_only=True, read_only=False)
        # BinaryIO
        return load_workbook(filename=source, data_only=True, read_only=False)

    @staticmethod
    def _dispatch(wb: Workbook) -> ParsedSoliqXltx:
        fmt = detect_format(wb)
        if fmt is SoliqXltxFormat.VAT_DECLARATION or fmt is SoliqXltxFormat.VAT_DECLARATION_V1:
            return parse_vat_declaration(wb)
        if fmt is SoliqXltxFormat.VAT_REGISTRY_ILOVA:
            return parse_vat_registry(wb)
        if fmt is SoliqXltxFormat.FORM_2_INCOME_STATEMENT:
            return parse_form2(wb)
        if fmt is SoliqXltxFormat.FORM_1_BALANCE_SHEET:
            return parse_form1(wb)
        raise UnsupportedFormatError(
            wb.sheetnames,
            f"format {fmt} not implemented yet "
            "(supports: VAT_DECLARATION, VAT_REGISTRY_ILOVA, "
            "FORM_2_INCOME_STATEMENT, FORM_1_BALANCE_SHEET)",
        )
