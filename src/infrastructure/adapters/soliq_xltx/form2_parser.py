"""Парсер «Отчёт о финансовых результатах» (Форма №2, 3 листа list01..list03).

Извлекает выручку, себестоимость, операционную прибыль, проценты, налог на
прибыль и чистую прибыль за отчётный YTD-период (column F/G) и за тот же
календарный период прошлого года (column D/E).

Multiplier для money-cells динамический из B24 list01 «Единица измерения»
(T2.1 CA-028): «тыс. сум.» → ×1000 (99% случаев), «млн. сум.» → ×1_000_000,
«сум.» → ×1 (полные). Unknown / empty → fallback ×1000 + warning.

Best-effort семантика (CA-014 hardening): cell-level проблема → warning + None.
Raise только на структурные: не тот формат (UnsupportedFormatError) или
отсутствие list02. Маркер ``'x'`` (Soliq «неприменимо») трактуется как None
без warning.

Signed-поля (net_profit, operating_profit, profit_before_tax) кодируются
парой колонок «Доходы / Расходы»: положительное значение в income column =
прибыль, в expense column = убыток. Парсер вычисляет ``income - expense`` с
'x'/None как 0; если **обе** ячейки пары пустые/x — возвращает None.

Координаты сверены с реальной выгрузкой Q4 2025 (ИНН 306399449 AZ RUHDIL):
- row 6  (010 Чистая выручка):       D6 prior, F6 current
- row 7  (020 Себестоимость):        E7 prior, G7 current
- row 8  (030 Валовая прибыль):      D8 prior, F8 current
- row 15 (100 Прибыль от основной):  (D15, E15) / (F15, G15) signed
- row 23 (180 Расходы в %):          E23 prior, G23 current
- row 29 (240 Прибыль до налога):    (D29, E29) / (F29, G29) signed
- row 30 (250 Налог на прибыль):     E30 prior, G30 current
- row 32 (270 Чистая прибыль):       (D32, E32) / (F32, G32) signed
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from decimal import Decimal
from typing import Any

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from domain.value_objects.money import Currency, Money
from infrastructure.adapters.soliq_xltx.errors import UnsupportedFormatError
from infrastructure.adapters.soliq_xltx.format_detector import (
    SoliqXltxFormat,
    detect_format,
)
from infrastructure.adapters.soliq_xltx.header_parser import (
    SoliqXltxHeader,
    parse_header,
    parse_unit_multiplier,
)

_logger = logging.getLogger(__name__)


@dataclass(slots=True)
class Form2IncomeStatementData:
    """Финансовые результаты за период YTD + прошлый год тот же период.

    Все поля Money — в UZS (парсер уже умножил на 1000 из тыс. сум.).

    ``current_period`` — period YTD as-of квартала из header (Q4 = annual).
    ``prior_year_period`` — тот же календарный период предыдущего года.

    Signed-поля (net_profit, operating_profit, profit_before_tax) могут быть
    отрицательными (убыток). Парные колонки «Доходы / Расходы» декодируются
    как ``income - expense``; если обе ячейки пусты — None.

    ``parse_warnings`` агрегирует header + body warnings.
    """

    header: SoliqXltxHeader
    revenue_current_period: Money | None
    revenue_prior_year_period: Money | None
    cost_of_sales_current: Money | None
    cost_of_sales_prior_year: Money | None
    gross_profit_current: Money | None
    gross_profit_prior_year: Money | None
    operating_profit_current: Money | None
    operating_profit_prior_year: Money | None
    interest_expense_current: Money | None
    interest_expense_prior_year: Money | None
    profit_before_tax_current: Money | None
    profit_before_tax_prior_year: Money | None
    profit_tax_current: Money | None
    profit_tax_prior_year: Money | None
    net_profit_current: Money | None
    net_profit_prior_year: Money | None
    parse_warnings: list[str] = field(default_factory=list)


def parse_form2(wb: Workbook) -> Form2IncomeStatementData:
    """Прочитать FORM_2 из openpyxl Workbook.

    Raise: ``UnsupportedFormatError`` если detect_format не вернул FORM_2 или
    отсутствует list02. Cell-level проблемы — warnings + None, не raise.
    """
    fmt = detect_format(wb)
    if fmt is not SoliqXltxFormat.FORM_2_INCOME_STATEMENT:
        raise UnsupportedFormatError(
            wb.sheetnames, f"expected FORM_2_INCOME_STATEMENT, got {fmt}"
        )

    header = parse_header(wb, fmt)
    if "list02" not in wb.sheetnames:
        raise UnsupportedFormatError(wb.sheetnames, "list02 missing in FORM_2")

    # T2.1 CA-028: multiplier динамический из B24 (раньше hardcoded ×1000).
    # Unknown / empty → fallback ×1000 + warning из helper'а (банк-friendly).
    multiplier, unit_warning = parse_unit_multiplier(wb, fmt)
    list02 = wb["list02"]
    warnings: list[str] = []
    if unit_warning is not None:
        warnings.append(unit_warning)

    return Form2IncomeStatementData(
        header=header,
        revenue_current_period=_money_kx(list02, "F6", warnings, multiplier),
        revenue_prior_year_period=_money_kx(list02, "D6", warnings, multiplier),
        cost_of_sales_current=_money_kx(list02, "G7", warnings, multiplier),
        cost_of_sales_prior_year=_money_kx(list02, "E7", warnings, multiplier),
        gross_profit_current=_money_kx(list02, "F8", warnings, multiplier),
        gross_profit_prior_year=_money_kx(list02, "D8", warnings, multiplier),
        operating_profit_current=_signed_money(list02, "F15", "G15", warnings, multiplier),
        operating_profit_prior_year=_signed_money(list02, "D15", "E15", warnings, multiplier),
        interest_expense_current=_money_kx(list02, "G23", warnings, multiplier),
        interest_expense_prior_year=_money_kx(list02, "E23", warnings, multiplier),
        profit_before_tax_current=_signed_money(list02, "F29", "G29", warnings, multiplier),
        profit_before_tax_prior_year=_signed_money(list02, "D29", "E29", warnings, multiplier),
        profit_tax_current=_money_kx(list02, "G30", warnings, multiplier),
        profit_tax_prior_year=_money_kx(list02, "E30", warnings, multiplier),
        net_profit_current=_signed_money(list02, "F32", "G32", warnings, multiplier),
        net_profit_prior_year=_signed_money(list02, "D32", "E32", warnings, multiplier),
        parse_warnings=[*header.parse_warnings, *warnings],
    )


def _money_kx(
    ws: Worksheet, coord: str, warnings: list[str], multiplier: Decimal
) -> Money | None:
    """Прочитать сумму в B23/B24-указанной единице → Money UZS (× multiplier).

    Best-effort. Multiplier приходит из ``parse_unit_multiplier`` —
    динамический (1 / 1_000 / 1_000_000) вместо hardcoded.
    """
    decimal_val = _to_decimal(ws[coord].value, ws.title, coord, warnings)
    if decimal_val is None:
        return None
    return Money(decimal_val * multiplier, Currency.UZS)


def _signed_money(
    ws: Worksheet,
    income_coord: str,
    expense_coord: str,
    warnings: list[str],
    multiplier: Decimal,
) -> Money | None:
    """Декодировать signed profit из пары (income, expense) колонок.

    ``income - expense`` × multiplier. 'x'/None → 0 для своей колонки. Если
    **обе** ячейки missing — возвращаем None (нет данных), не Money(0).
    """
    income_raw = ws[income_coord].value
    expense_raw = ws[expense_coord].value

    income_missing = _is_missing(income_raw)
    expense_missing = _is_missing(expense_raw)
    if income_missing and expense_missing:
        return None

    income = (
        Decimal(0)
        if income_missing
        else (_to_decimal(income_raw, ws.title, income_coord, warnings) or Decimal(0))
    )
    expense = (
        Decimal(0)
        if expense_missing
        else (_to_decimal(expense_raw, ws.title, expense_coord, warnings) or Decimal(0))
    )
    return Money((income - expense) * multiplier, Currency.UZS)


def _is_missing(raw: Any) -> bool:
    """Cell считается отсутствующим если None или 'x'/'Х' marker / пустая строка.

    Маркер ``'x'`` (латинский) и ``'Х'`` (кириллический) — стандартное Soliq
    обозначение «поле неприменимо», не ошибка → молча → None без warning.
    """
    if raw is None:
        return True
    return isinstance(raw, str) and raw.strip().lower() in ("x", "х", "")


def _to_decimal(
    raw: Any, sheet: str, coord: str, warnings: list[str]
) -> Decimal | None:
    """Best-effort numeric → Decimal. 'x'/None молчаливое None; мусор → warning + None."""
    if _is_missing(raw):
        return None
    if isinstance(raw, bool):
        # bool — подкласс int, отсечь до проверки int/float
        _warn(warnings, sheet, coord, f"unsupported bool: {raw!r}")
        return None
    if isinstance(raw, (int, float)):
        return Decimal(str(raw))
    if isinstance(raw, str):
        try:
            return Decimal(raw.strip().replace(",", "."))
        except Exception:
            _warn(warnings, sheet, coord, f"non-numeric money: {raw!r}")
            return None
    _warn(warnings, sheet, coord, f"unsupported type: {type(raw).__name__}")
    return None


def _warn(warnings: list[str], sheet: str, coord: str, reason: str) -> None:
    msg = f"{sheet}!{coord}: {reason}"
    warnings.append(msg)
    _logger.warning(
        "xltx.form2_cell_skipped sheet=%s coord=%s reason=%s",
        sheet,
        coord,
        reason,
    )
