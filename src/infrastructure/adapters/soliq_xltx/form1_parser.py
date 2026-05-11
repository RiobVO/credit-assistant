"""Парсер «Бухгалтерский баланс» (Форма №1, list01..list04).

Извлекает 12 балансовых показателей × 2 столбца (period_end / period_start)
из единственного содержательного листа ``list02``. Остальные листы (list03
забалансовые ценности, list04 расшифровка просроченной задолженности) для
скоринга не используются.

Все суммы в файле — в **тыс. сум** (B23 list01 «Единица измерения, тыс. сум.»).
Парсер умножает на 1000 и возвращает Money в UZS.

Best-effort семантика (CA-014): cell-level проблема → warning + None. Raise
только на структурные: не тот формат (UnsupportedFormatError) или отсутствие
list02. Маркер ``'x'``/``'Х'`` (Soliq «неприменимо») — None без warning.

Семантика колонок отличается от FORM_2 (см. CA-029):
- column **E** = «На конец отчётного периода» → ``*_period_end`` (текущая дата)
- column **D** = «На начало отчётного периода» → ``*_period_start`` (опен)

Это balance sheet — точечные срезы на две даты в рамках одного отчётного
периода, а не сравнение с прошлым годом.

Координаты сверены с реальной выгрузкой Q4 2025 (ИНН 201308534 QADR DON NON):
- row 10 (стр. 012 — Остаточная стоимость ОС):       fixed_assets
- row 25 (стр. 130 — Итого по разделу I актива):     non_current_assets
- row 27 (стр. 140 — Товарно-материальные запасы):   inventory
- row 34 (стр. 210 — Дебиторы всего):                receivables
- row 46 (стр. 320 — Денежные средства всего):       cash_and_equivalents
- row 53 (стр. 390 — Итого по разделу II актива):    current_assets
- row 54 (стр. 400 — Всего по активу):               total_assets
- row 64 (стр. 480 — Итого по разделу I пассива):    equity
- row 66 (стр. 490 — Долгосрочные обязательства):    long_term_liabilities
- row 78 (стр. 600 — Текущие обязательства):         short_term_liabilities
- row 97 (стр. 770 — Итого по разделу II пассива):   total_liabilities

``total_debt`` — derived агрегат пяти долговых строк (570+580+730+740+750):
долгосрочные банковские кредиты, долгосрочные займы, краткосрочные банковские
кредиты, краткосрочные займы, текущая часть долгосрочных. Правило: хотя бы
одна не-None → агрегат не None; все None → None.

Balance equation sanity (warning, не raise): для каждого периода проверяется
``total_assets ≈ equity + total_liabilities`` с допуском 0.5%. Расхождение
больше — запись в ``parse_warnings`` с конкретными цифрами.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from decimal import Decimal
from typing import Any

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from domain.value_objects.money import Currency, Money
from infrastructure.adapters.soliq_xltx.errors import UnsupportedFormatError
from infrastructure.adapters.soliq_xltx.format_detector import (
    SoliqXltxFormat,
    detect_format,
)
from infrastructure.adapters.soliq_xltx.header_parser import (
    SoliqXltxHeader,
    parse_header,
)

_logger = logging.getLogger(__name__)

# Все money-cells form1 — в тыс. сум; домен оперирует UZS, поэтому ×1000.
# TODO[CA-028]: динамическая детекция ЕИ из B23 list01 (сейчас hardcoded).
_THOUSANDS_MULTIPLIER = Decimal(1000)

# Допуск balance-equation: 0.5% (5‰). Реальные балансы дают округления на
# уровне единиц тыс. сум — 1‰ ложно фейлил бы. 5‰ покрывает rounding + ловит
# реальные ошибки переноса.
_BALANCE_TOLERANCE = Decimal("0.005")

# Координаты пяти долговых строк, агрегирующихся в total_debt.
# (label, period_end_coord, period_start_coord)
_DEBT_COMPONENT_CELLS: tuple[tuple[str, str, str], ...] = (
    ("570 LT bank loans", "E75", "D75"),
    ("580 LT borrowings", "E76", "D76"),
    ("730 ST bank loans", "E93", "D93"),
    ("740 ST borrowings", "E94", "D94"),
    ("750 current portion of LT", "E95", "D95"),
)


@dataclass(slots=True)
class Form1BalanceSheetData:
    """Балансовый срез на конец и начало отчётного периода.

    Все Money — в UZS (×1000 из тыс. сум.). ``*_period_end`` соответствует
    column E xltx («На конец отчётного периода»), ``*_period_start`` — column D
    («На начало отчётного периода»). Это две точки на оси времени в рамках
    одного отчёта, не сравнение с прошлым годом.

    ``total_debt_*`` — derived sum из 5 долговых строк (см. модульный docstring).

    ``parse_warnings`` агрегирует header + cell-level warnings + balance-
    equation sanity.
    """

    header: SoliqXltxHeader
    fixed_assets_period_end: Money | None
    fixed_assets_period_start: Money | None
    non_current_assets_period_end: Money | None
    non_current_assets_period_start: Money | None
    inventory_period_end: Money | None
    inventory_period_start: Money | None
    receivables_period_end: Money | None
    receivables_period_start: Money | None
    cash_and_equivalents_period_end: Money | None
    cash_and_equivalents_period_start: Money | None
    current_assets_period_end: Money | None
    current_assets_period_start: Money | None
    total_assets_period_end: Money | None
    total_assets_period_start: Money | None
    equity_period_end: Money | None
    equity_period_start: Money | None
    long_term_liabilities_period_end: Money | None
    long_term_liabilities_period_start: Money | None
    short_term_liabilities_period_end: Money | None
    short_term_liabilities_period_start: Money | None
    total_liabilities_period_end: Money | None
    total_liabilities_period_start: Money | None
    total_debt_period_end: Money | None
    total_debt_period_start: Money | None
    parse_warnings: list[str] = field(default_factory=list)


def parse_form1(wb: Workbook) -> Form1BalanceSheetData:
    """Прочитать FORM_1 из openpyxl Workbook.

    Raise ``UnsupportedFormatError``, если detect_format вернул не FORM_1
    или нет list02. Cell-level проблемы — warnings + None, не raise.
    """
    fmt = detect_format(wb)
    if fmt is not SoliqXltxFormat.FORM_1_BALANCE_SHEET:
        raise UnsupportedFormatError(
            wb.sheetnames, f"expected FORM_1_BALANCE_SHEET, got {fmt}"
        )

    header = parse_header(wb, fmt)
    if "list02" not in wb.sheetnames:
        raise UnsupportedFormatError(wb.sheetnames, "list02 missing in FORM_1")

    list02 = wb["list02"]
    warnings: list[str] = []

    fixed_end = _money_kx(list02, "E10", warnings)
    fixed_start = _money_kx(list02, "D10", warnings)
    non_current_end = _money_kx(list02, "E25", warnings)
    non_current_start = _money_kx(list02, "D25", warnings)
    inventory_end = _money_kx(list02, "E27", warnings)
    inventory_start = _money_kx(list02, "D27", warnings)
    receivables_end = _money_kx(list02, "E34", warnings)
    receivables_start = _money_kx(list02, "D34", warnings)
    cash_end = _money_kx(list02, "E46", warnings)
    cash_start = _money_kx(list02, "D46", warnings)
    current_end = _money_kx(list02, "E53", warnings)
    current_start = _money_kx(list02, "D53", warnings)
    total_assets_end = _money_kx(list02, "E54", warnings)
    total_assets_start = _money_kx(list02, "D54", warnings)
    equity_end = _money_kx(list02, "E64", warnings)
    equity_start = _money_kx(list02, "D64", warnings)
    lt_liab_end = _money_kx(list02, "E66", warnings)
    lt_liab_start = _money_kx(list02, "D66", warnings)
    st_liab_end = _money_kx(list02, "E78", warnings)
    st_liab_start = _money_kx(list02, "D78", warnings)
    total_liab_end = _money_kx(list02, "E97", warnings)
    total_liab_start = _money_kx(list02, "D97", warnings)

    debt_end = _aggregate_debt(list02, "period_end", warnings)
    debt_start = _aggregate_debt(list02, "period_start", warnings)

    _check_balance_equation(
        "period_end", total_assets_end, equity_end, total_liab_end, warnings
    )
    _check_balance_equation(
        "period_start", total_assets_start, equity_start, total_liab_start, warnings
    )

    return Form1BalanceSheetData(
        header=header,
        fixed_assets_period_end=fixed_end,
        fixed_assets_period_start=fixed_start,
        non_current_assets_period_end=non_current_end,
        non_current_assets_period_start=non_current_start,
        inventory_period_end=inventory_end,
        inventory_period_start=inventory_start,
        receivables_period_end=receivables_end,
        receivables_period_start=receivables_start,
        cash_and_equivalents_period_end=cash_end,
        cash_and_equivalents_period_start=cash_start,
        current_assets_period_end=current_end,
        current_assets_period_start=current_start,
        total_assets_period_end=total_assets_end,
        total_assets_period_start=total_assets_start,
        equity_period_end=equity_end,
        equity_period_start=equity_start,
        long_term_liabilities_period_end=lt_liab_end,
        long_term_liabilities_period_start=lt_liab_start,
        short_term_liabilities_period_end=st_liab_end,
        short_term_liabilities_period_start=st_liab_start,
        total_liabilities_period_end=total_liab_end,
        total_liabilities_period_start=total_liab_start,
        total_debt_period_end=debt_end,
        total_debt_period_start=debt_start,
        parse_warnings=[*header.parse_warnings, *warnings],
    )


def _aggregate_debt(
    ws: Worksheet, column: str, warnings: list[str]
) -> Money | None:
    """Сумма 5 долговых строк (570+580+730+740+750) в одном периоде.

    Правило: хотя бы одна ячейка не-None → агрегат = сумма не-None. Все None
    → None (нет данных по долгу). Тихий 'x'/missing не считается «есть данные»
    — он трактуется как 0 в составе агрегата, но сам по себе не активирует
    агрегат.
    """
    column_index = 1 if column == "period_end" else 2  # 1-based: E=1st, D=2nd
    values: list[Decimal] = []
    has_value = False
    for _label, end_coord, start_coord in _DEBT_COMPONENT_CELLS:
        coord = end_coord if column_index == 1 else start_coord
        raw = ws[coord].value
        if _is_missing(raw):
            continue
        decimal_val = _to_decimal(raw, ws.title, coord, warnings)
        if decimal_val is None:
            continue
        values.append(decimal_val)
        has_value = True
    if not has_value:
        return None
    total = sum(values, Decimal(0))
    return Money(total * _THOUSANDS_MULTIPLIER, Currency.UZS)


def _check_balance_equation(
    label: str,
    total_assets: Money | None,
    equity: Money | None,
    total_liabilities: Money | None,
    warnings: list[str],
) -> None:
    """Sanity: total_assets ≈ equity + total_liabilities с допуском 0.5%.

    Skip если любая компонента None — недостаточно данных, парсер не делает
    выводов. Расхождение > 0.5% от |total_assets| → warning с цифрами.
    """
    if total_assets is None or equity is None or total_liabilities is None:
        return
    a = total_assets.amount
    el = equity.amount + total_liabilities.amount
    delta = abs(a - el)
    if a == 0:
        # Особый случай: нулевой баланс. Любое расхождение значимо в абсолюте,
        # но допускаем равные нули как «нет данных», без warning.
        if delta == 0:
            return
        warnings.append(
            f"balance_equation_mismatch ({label}): "
            f"A=0, E+L={el}, delta={delta} (assets zero, components non-zero)"
        )
        return
    ratio = delta / abs(a)
    if ratio > _BALANCE_TOLERANCE:
        pct = ratio * Decimal(100)
        warnings.append(
            f"balance_equation_mismatch ({label}): "
            f"A={a}, E+L={el}, delta={pct:.3f}%"
        )


def _money_kx(ws: Worksheet, coord: str, warnings: list[str]) -> Money | None:
    """Прочитать сумму в тыс. сум → Money UZS (× 1000). Best-effort."""
    decimal_val = _to_decimal(ws[coord].value, ws.title, coord, warnings)
    if decimal_val is None:
        return None
    return Money(decimal_val * _THOUSANDS_MULTIPLIER, Currency.UZS)


def _is_missing(raw: Any) -> bool:
    """Cell считается отсутствующим: None / 'x' / 'Х' / пустая строка."""
    if raw is None:
        return True
    return isinstance(raw, str) and raw.strip().lower() in ("x", "х", "")


def _to_decimal(
    raw: Any, sheet: str, coord: str, warnings: list[str]
) -> Decimal | None:
    """Best-effort numeric → Decimal. 'x'/None молчаливое None; мусор → warning + None."""
    if _is_missing(raw):
        return None
    if isinstance(raw, bool):
        _warn(warnings, sheet, coord, f"unsupported bool: {raw!r}")
        return None
    if isinstance(raw, (int, float)):
        return Decimal(str(raw))
    if isinstance(raw, str):
        try:
            return Decimal(raw.strip().replace(",", "."))
        except Exception:
            _warn(warnings, sheet, coord, f"non-numeric money: {raw!r}")
            return None
    _warn(warnings, sheet, coord, f"unsupported type: {type(raw).__name__}")
    return None


def _warn(warnings: list[str], sheet: str, coord: str, reason: str) -> None:
    msg = f"{sheet}!{coord}: {reason}"
    warnings.append(msg)
    _logger.warning(
        "xltx.form1_cell_skipped sheet=%s coord=%s reason=%s",
        sheet,
        coord,
        reason,
    )
