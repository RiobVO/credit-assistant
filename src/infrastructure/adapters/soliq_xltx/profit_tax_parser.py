"""Парсер «Расчёт налога на прибыль юридических лиц» (PROFIT_TAX, 15-16 листов).

Сводка вычислений расположена на list01 rows 28-46: column K = код строки,
column L = значение. Парсер читает только two key cells (минимальный scope
T2.3, остальные поля extend по требованию):

- L31 (код 030 «Налогооблагаемая прибыль») — signed, может быть отрицательной
  (убыток). Money | None в DTO.
- L39 (код 080 «Сумма налога на прибыль – всего») — gross computed, идёт в
  `ParsedFinancials.taxes_paid_by_year` через use-case wiring.

Суммы в **полных сум** (не тыс. сум как в FORM_2) — multiplier ×1.
Cross-check: PROFIT_TAX L29 = 1,116,265,183 сум ≈ FORM_2 F6 × 1000 =
1,116,265,000 сум для ZAMIN NOZ NEMATLARI (305002665).

Best-effort семантика (CA-014): cell-level проблема → warning + None.
Raise только структурные: не тот формат (UnsupportedFormatError) или
отсутствие list01. Маркер ``'x'`` (Soliq «неприменимо») и ``None`` — молча
None без warning.

Координаты сверены с реальными выгрузками Q4 2025 от 4 фирм
(profit_tax_2025_{201308534,305002665,305738460,308747266}_full.xltx).
Quarterly version (Q1/Q2/Q3) показывает дрейф L36 (ставка) и других ячеек —
use-case `parse_manual_input_files` skip'ит quarterly PROFIT_TAX отдельно.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from decimal import Decimal
from typing import Any

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from domain.value_objects.money import Currency, Money
from infrastructure.adapters.soliq_xltx.errors import UnsupportedFormatError
from infrastructure.adapters.soliq_xltx.format_detector import (
    SoliqXltxFormat,
    detect_format,
)
from infrastructure.adapters.soliq_xltx.header_parser import (
    SoliqXltxHeader,
    parse_header,
)

_logger = logging.getLogger(__name__)


@dataclass(slots=True)
class ProfitTaxData:
    """Расчёт налога на прибыль: ключевые поля для wiring в `taxes_paid_by_year`.

    Все Money — в UZS (полные сум, multiplier ×1).

    ``taxable_profit`` — L31 (код 030), signed. Отрицательное Money = убыток.
    ``profit_tax_total`` — L39 (код 080), gross computed; всегда non-negative
    в Soliq-реальности (нулевой при убытке).

    ``parse_warnings`` агрегирует header + body warnings (best-effort).
    """

    header: SoliqXltxHeader
    taxable_profit: Money | None
    profit_tax_total: Money | None
    parse_warnings: list[str] = field(default_factory=list)


def parse_profit_tax(wb: Workbook) -> ProfitTaxData:
    """Прочитать PROFIT_TAX из openpyxl Workbook.

    Raise:
    - ``UnsupportedFormatError`` если detect_format не вернул PROFIT_TAX
      или отсутствует list01 (структурные ошибки).

    Cell-level проблемы — warnings + None, не raise.
    """
    fmt = detect_format(wb)
    if fmt is not SoliqXltxFormat.PROFIT_TAX:
        raise UnsupportedFormatError(
            wb.sheetnames, f"expected PROFIT_TAX, got {fmt}"
        )

    if "list01" not in wb.sheetnames:
        raise UnsupportedFormatError(wb.sheetnames, "list01 missing in PROFIT_TAX")

    header = parse_header(wb, fmt)
    list01 = wb["list01"]
    warnings: list[str] = []

    return ProfitTaxData(
        header=header,
        taxable_profit=_money_full(list01, "L31", warnings),
        profit_tax_total=_money_full(list01, "L39", warnings),
        parse_warnings=[*header.parse_warnings, *warnings],
    )


def _money_full(ws: Worksheet, coord: str, warnings: list[str]) -> Money | None:
    """Прочитать сумму (×1, полные сум) → Money UZS. Best-effort."""
    decimal_val = _to_decimal(ws[coord].value, ws.title, coord, warnings)
    if decimal_val is None:
        return None
    return Money(decimal_val, Currency.UZS)


def _is_missing(raw: Any) -> bool:
    """Cell считается отсутствующим если None или 'x'/'Х' marker / пустая строка.

    Маркер ``'x'`` (латинский) и ``'Х'`` (кириллический) — стандартное Soliq
    обозначение «поле неприменимо», не ошибка → молча None без warning.
    """
    if raw is None:
        return True
    return isinstance(raw, str) and raw.strip().lower() in ("x", "х", "")


def _to_decimal(
    raw: Any, sheet: str, coord: str, warnings: list[str]
) -> Decimal | None:
    """Best-effort numeric → Decimal. 'x'/None молчаливое None; мусор → warning + None."""
    if _is_missing(raw):
        return None
    if isinstance(raw, bool):
        # bool — подкласс int, отсечь до проверки int/float
        _warn(warnings, sheet, coord, f"unsupported bool: {raw!r}")
        return None
    if isinstance(raw, (int, float)):
        return Decimal(str(raw))
    if isinstance(raw, str):
        try:
            return Decimal(raw.strip().replace(",", "."))
        except Exception:
            _warn(warnings, sheet, coord, f"non-numeric money: {raw!r}")
            return None
    _warn(warnings, sheet, coord, f"unsupported type: {type(raw).__name__}")
    return None


def _warn(warnings: list[str], sheet: str, coord: str, reason: str) -> None:
    msg = f"{sheet}!{coord}: {reason}"
    warnings.append(msg)
    _logger.warning(
        "xltx.profit_tax_cell_skipped sheet=%s coord=%s reason=%s",
        sheet,
        coord,
        reason,
    )
