"""Парсер «Расчёт НДС» (декларация НДС, 8 листов list01..list08).

Извлекает декларированные обороты по реализации с разбивкой по каналам
(ЭСФ / ККМ / экспорт / marketplace / прочие) и сумму НДС, начисленную и
подлежащую зачёту.

Best-effort семантика (CA-014 hardening): любая cell-level ошибка
(garbage в money-cell, неожиданный тип) → warning + None. Структурные
ошибки (отсутствие list02/list04) по-прежнему raise — без этих листов
парсер не знает откуда читать.

Координаты сверены с реальной выгрузкой папы (март 2026):
- list02 (Приложение №1): R6 итого по реализации (стр.010), R7-R11 разбивка
  по каналам (стр.0101..0105). Колонки: F стоимость, G сумма НДС.
- list04 (Приложение №3): R7 НДС к зачёту (нарастающий с начала года, стр.010),
  R37 НДС подлежащий зачёту (стр.040, col G).

Ячейки с символом 'x' (экспорт без НДС) → ``None``.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from decimal import Decimal
from typing import Any

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from domain.value_objects.money import Currency, Money
from infrastructure.adapters.soliq_xltx.errors import UnsupportedFormatError
from infrastructure.adapters.soliq_xltx.format_detector import (
    SoliqXltxFormat,
    detect_format,
)
from infrastructure.adapters.soliq_xltx.header_parser import (
    SoliqXltxHeader,
    parse_header,
)

_logger = logging.getLogger(__name__)


@dataclass(slots=True)
class VatDeclarationData:
    """Декларированные данные НДС за налоговый период.

    ``vat_charged_total`` — сумма НДС, начисленная за период (стр.010 list02 G6).
    Это же значение пойдёт в ``FinancialReport.vat_declared`` при сборке snapshot
    в Day 2.

    ``sales_via_*`` — стоимость без НДС по каналам реализации (стр.0101-0105 col F).
    ``*_vat`` — соответствующая сумма НДС (col G); может быть ``None`` для каналов,
    где Soliq ставит 'x' (например, экспорт по нулевой ставке).

    ``parse_warnings`` агрегирует cell-level проблемы из header + body.
    """

    header: SoliqXltxHeader
    sales_total_excl_vat: Money | None
    vat_charged_total: Money | None
    sales_via_esf: Money | None
    vat_via_esf: Money | None
    sales_via_kkm: Money | None
    vat_via_kkm: Money | None
    sales_via_export: Money | None
    vat_via_export: Money | None
    sales_via_marketplace: Money | None
    vat_via_marketplace: Money | None
    sales_via_other: Money | None
    vat_via_other: Money | None
    vat_to_offset_year_cumulative: Money | None
    vat_to_offset_total: Money | None
    parse_warnings: list[str] = field(default_factory=list)


def parse_vat_declaration(wb: Workbook) -> VatDeclarationData:
    """Прочитать VAT-декларацию из openpyxl Workbook.

    Бросает ``UnsupportedFormatError``, если формат не VAT_DECLARATION либо
    отсутствуют list02/list04.
    """
    fmt = detect_format(wb)
    if fmt is not SoliqXltxFormat.VAT_DECLARATION:
        raise UnsupportedFormatError(wb.sheetnames, f"expected VAT_DECLARATION, got {fmt}")

    header = parse_header(wb, fmt)
    if "list02" not in wb.sheetnames:
        raise UnsupportedFormatError(wb.sheetnames, "list02 missing in VAT declaration")
    if "list04" not in wb.sheetnames:
        raise UnsupportedFormatError(wb.sheetnames, "list04 missing in VAT declaration")

    list02 = wb["list02"]
    list04 = wb["list04"]
    body_warnings: list[str] = []

    return VatDeclarationData(
        header=header,
        sales_total_excl_vat=_money(list02, "F6", body_warnings),
        vat_charged_total=_money(list02, "G6", body_warnings),
        sales_via_esf=_money(list02, "F7", body_warnings),
        vat_via_esf=_money(list02, "G7", body_warnings),
        sales_via_kkm=_money(list02, "F8", body_warnings),
        vat_via_kkm=_money(list02, "G8", body_warnings),
        sales_via_export=_money(list02, "F9", body_warnings),
        vat_via_export=_money(list02, "G9", body_warnings),
        sales_via_marketplace=_money(list02, "F10", body_warnings),
        vat_via_marketplace=_money(list02, "G10", body_warnings),
        sales_via_other=_money(list02, "F11", body_warnings),
        vat_via_other=_money(list02, "G11", body_warnings),
        vat_to_offset_year_cumulative=_money(list04, "G7", body_warnings),
        vat_to_offset_total=_money(list04, "G37", body_warnings),
        parse_warnings=[*header.parse_warnings, *body_warnings],
    )


def _money(ws: Worksheet, coord: str, warnings: list[str]) -> Money | None:
    """Прочитать денежную сумму из ячейки.

    Возвращает ``None`` для пустых cells и для ячеек со значением 'x'/'Х'
    (Soliq использует X-маркер на полях, где значение неприменимо). При
    невалидном содержимом — warning + None (не raise).
    """
    raw = ws[coord].value
    if raw is None:
        return None
    if isinstance(raw, str):
        text = raw.strip()
        if text.lower() in ("x", "х", ""):  # Латинский x и кириллический Х
            return None
        try:
            return Money(Decimal(text.replace(",", ".")), Currency.UZS)
        except Exception:
            _warn(warnings, ws.title, coord, f"non-numeric money: {raw!r}")
            return None
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        return Money(_decimal_from_number(raw), Currency.UZS)
    _warn(warnings, ws.title, coord, f"unsupported money type: {type(raw).__name__}")
    return None


def _decimal_from_number(raw: Any) -> Decimal:
    """Конверсия numeric cell value в Decimal через str для сохранения точности."""
    return Decimal(str(raw))


def _warn(warnings: list[str], sheet: str, coord: str, reason: str) -> None:
    msg = f"{sheet}!{coord}: {reason}"
    warnings.append(msg)
    _logger.warning(
        "xltx.declaration_cell_skipped sheet=%s coord=%s reason=%s",
        sheet,
        coord,
        reason,
    )
