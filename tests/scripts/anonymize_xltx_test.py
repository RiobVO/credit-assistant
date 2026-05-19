"""Unit-тесты для scripts.anonymize_xltx — synthetic xlsx, без зависимости от fixtures."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from scripts.anonymize_xltx import (
    _anonymize_filename,
    _deterministic_inn,
    anonymize_xltx,
)

_CellValue = str | int | float | None


def _make_xlsx(path: Path, cells: dict[tuple[int, int], _CellValue]) -> None:
    """Соберёт synthetic xlsx с заданными значениями ячеек."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    for (row, col), value in cells.items():
        ws.cell(row=row, column=col, value=value)
    wb.save(str(path))
    wb.close()


def _read_cells(
    path: Path, coords: list[tuple[int, int]]
) -> dict[tuple[int, int], object]:
    wb = load_workbook(filename=str(path), data_only=False)
    try:
        ws = wb.active
        assert ws is not None
        return {(r, c): ws.cell(row=r, column=c).value for (r, c) in coords}
    finally:
        wb.close()


def test_inn_hash_deterministic() -> None:
    """ИНН → hash детерминирован и сохраняет длину."""
    a1 = _deterministic_inn("201308534")
    a2 = _deterministic_inn("201308534")
    assert a1 == a2
    assert len(a1) == 9
    assert a1 != "201308534"

    long1 = _deterministic_inn("12345678901234")
    long2 = _deterministic_inn("12345678901234")
    assert long1 == long2
    assert len(long1) == 14

    # Разные input → разный output (collision-resistance baseline)
    assert _deterministic_inn("201308534") != _deterministic_inn("305002665")


def test_formula_preserved(tmp_path: Path) -> None:
    """Cell с формулой не трогается — data_type='f' остаётся."""
    src = tmp_path / "src.xltx"
    dst = tmp_path / "dst.xltx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.cell(row=1, column=1, value=50000)
    ws.cell(row=1, column=2, value=70000)
    ws.cell(row=1, column=3, value="=A1+B1")
    wb.save(str(src))
    wb.close()

    anonymize_xltx(src, dst)

    wb2 = load_workbook(filename=str(dst), data_only=False)
    try:
        ws2 = wb2.active
        assert ws2 is not None
        formula_cell = ws2.cell(row=1, column=3)
        assert formula_cell.data_type == "f"
        assert formula_cell.value == "=A1+B1"
    finally:
        wb2.close()


def test_sentinel_strings_untouched(tmp_path: Path) -> None:
    """Format-sentinel строки не содержат INN/company-markers → unchanged."""
    src = tmp_path / "src.xltx"
    dst = tmp_path / "dst.xltx"
    _make_xlsx(
        src,
        {
            (1, 1): "Расчёт НДС",
            (2, 1): "Бухгалтерский баланс",
            (3, 1): "Отчёт о финансовых результатах",
            (4, 1): "Приложение №4",
        },
    )
    anonymize_xltx(src, dst)
    cells = _read_cells(dst, [(1, 1), (2, 1), (3, 1), (4, 1)])
    assert cells[(1, 1)] == "Расчёт НДС"
    assert cells[(2, 1)] == "Бухгалтерский баланс"
    assert cells[(3, 1)] == "Отчёт о финансовых результатах"
    assert cells[(4, 1)] == "Приложение №4"


def test_small_numeric_untouched(tmp_path: Path) -> None:
    """Numeric ≤ AMOUNT_THRESHOLD и year-like values остаются как есть."""
    src = tmp_path / "src.xltx"
    dst = tmp_path / "dst.xltx"
    _make_xlsx(
        src,
        {
            (1, 1): 500,
            (2, 1): 9999,
            (3, 1): 2025,
            (4, 1): 1900,
        },
    )
    anonymize_xltx(src, dst)
    cells = _read_cells(dst, [(1, 1), (2, 1), (3, 1), (4, 1)])
    assert cells[(1, 1)] == 500
    assert cells[(2, 1)] == 9999
    assert cells[(3, 1)] == 2025
    assert cells[(4, 1)] == 1900


def test_company_name_replaced(tmp_path: Path) -> None:
    """Cells с ООО/ОАО → ОБЕЗЛИЧЕНО NNN, per-file counter."""
    src = tmp_path / "src.xltx"
    dst = tmp_path / "dst.xltx"
    _make_xlsx(
        src,
        {
            (1, 1): "ООО РОМАШКА",
            (2, 1): "ОАО ИНТЕГРАЛ",
            (3, 1): "Просто строка",
        },
    )
    anonymize_xltx(src, dst)
    cells = _read_cells(dst, [(1, 1), (2, 1), (3, 1)])
    assert cells[(1, 1)] == "ОБЕЗЛИЧЕНО 001"
    assert cells[(2, 1)] == "ОБЕЗЛИЧЕНО 002"
    assert cells[(3, 1)] == "Просто строка"


def test_anonymize_filename_substitutes_inn() -> None:
    """Filename helper: INN в стеме → anon-hash, _full → _anon."""
    p1 = Path("form1_2025_201308534_full.xltx")
    result = _anonymize_filename(p1)
    assert result.endswith("_anon.xltx")
    assert "201308534" not in result
    assert result.startswith("form1_2025_")

    p2 = Path("form1_q4_2025_full.xltx")
    assert _anonymize_filename(p2) == "form1_q4_2025_anon.xltx"

    same_inn_form1 = _anonymize_filename(Path("form1_2025_201308534_full.xltx"))
    same_inn_form2 = _anonymize_filename(Path("form2_2025_201308534_full.xltx"))
    assert same_inn_form1.split("_")[2] == same_inn_form2.split("_")[2]
