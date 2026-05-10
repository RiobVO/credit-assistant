"""In-memory builders openpyxl Workbook'ов под форматы выгрузок my3.soliq.uz.

Тесты адаптера ``soliq_xltx`` используют эти factories, чтобы не хранить
.xltx-бинарники в репозитории. Координаты ячеек сверены с реальными выгрузками
(папа, ИНН 306399449), мартовская декларация НДС + ilova-реестр.

Factories укладывают **минимально необходимый** набор cells — то, что читает парсер.
Не нужно воссоздавать всю форму с её десятками заголовков-описаний.
"""

from __future__ import annotations

from collections.abc import Sequence
from typing import Any

from openpyxl import Workbook
from openpyxl.workbook.workbook import Workbook as WorkbookT


def build_vat_declaration_wb(
    *,
    inn: str | float = 306399449,
    organization_name: str = '"AZ RUHDIL SAVDO" MAS\'ULIYATI CHEKLANGAN JAMIYAT',
    period_year: int = 2026,
    period_kind: str = "месяц",
    submitted_at: str = "20.04.2026",
    sales_total_excl_vat: float | None = 523333214.06,
    sales_total_vat: float | None = 62799985.69,
    sales_via_esf: tuple[float | None, float | None] = (523333214.06, 62799985.94),
    sales_via_kkm: tuple[float | None, float | None] = (0.0, 0.0),
    sales_via_export: tuple[float | None, Any] = (0.0, "x"),
    sales_via_marketplace: tuple[float | None, float | None] = (0.0, 0.0),
    sales_via_other: tuple[float | None, float | None] = (None, 0.0),
    vat_to_offset_year_cumulative: tuple[float | None, float | None] = (
        1416681390.98,
        169838602.88,
    ),
    vat_to_offset_total: float | None = 167588345.73,
) -> WorkbookT:
    """Декларация НДС (Расчёт НДС, 8 листов: list01..list08).

    Только ключевые координаты, читаемые парсером:
    - list01: D3 ИНН, H9 организация, K5 period_kind, O5 period_year, H17 submitted_at,
      D8 sentinel-фраза для format_detector.
    - list02 (Приложение №1, обороты по реализации): R6 итого / R7 ЭСФ / R8 ККМ /
      R9 экспорт / R10 marketplace / R11 прочие, столбцы F стоимость, G НДС.
    - list04 (Приложение №3, НДС к зачёту): R7 cumulative с начала года, R37 итого
      подлежит зачёту (col G).
    Остальные листы list03/05-08 — пустые, нужны только чтобы format_detector
    видел 8 листов.
    """
    wb = Workbook()
    ws1 = wb.active
    assert ws1 is not None
    ws1.title = "list01"

    # header coordinates
    ws1["B3"] = "ИНН"
    ws1["D3"] = inn
    ws1["C5"] = "Вид документа"
    ws1["E5"] = 1.0
    ws1["F5"] = "Ҳисобот"
    ws1["G5"] = "Отчетный период"
    ws1["K5"] = period_kind
    ws1["M5"] = "Налоговый период"
    ws1["O5"] = float(period_year)
    ws1["P5"] = "год"
    # sentinel-cell для format_detector
    ws1["D8"] = "СВЕДЕНИЯ о плательщике налога на добавленную стоимость"
    ws1["B9"] = "Полное наименование налогоплательщика"
    ws1["H9"] = organization_name
    ws1["B17"] = "Срок представления Расчета (день/месяц/год)"
    ws1["H17"] = submitted_at

    ws2 = wb.create_sheet("list02")
    ws2["B4"] = "Показатели"
    ws2["E4"] = "Код строки"
    ws2["F4"] = "Стоимость"
    ws2["G4"] = "Сумма НДС"
    # R6 = str. 010 итого по реализации
    ws2["E6"] = "010"
    ws2["F6"] = sales_total_excl_vat
    ws2["G6"] = sales_total_vat
    # R7 = 0101 ЭСФ
    ws2["E7"] = "0101"
    ws2["F7"] = sales_via_esf[0]
    ws2["G7"] = sales_via_esf[1]
    # R8 = 0102 ККМ
    ws2["E8"] = "0102"
    ws2["F8"] = sales_via_kkm[0]
    ws2["G8"] = sales_via_kkm[1]
    # R9 = 0103 экспорт (НДС обычно 'x')
    ws2["E9"] = "0103"
    ws2["F9"] = sales_via_export[0]
    ws2["G9"] = sales_via_export[1]
    # R10 = 0104 marketplace
    ws2["E10"] = "0104"
    ws2["F10"] = sales_via_marketplace[0]
    ws2["G10"] = sales_via_marketplace[1]
    # R11 = 0105 прочие
    ws2["E11"] = "0105"
    ws2["F11"] = sales_via_other[0]
    ws2["G11"] = sales_via_other[1]

    wb.create_sheet("list03")  # пустой

    ws4 = wb.create_sheet("list04")
    ws4["B5"] = "Показатели"
    ws4["E5"] = "Код строки"
    ws4["F5"] = "Стоимость"
    ws4["G5"] = "Сумма НДС"
    ws4["E7"] = "010"
    ws4["F7"] = vat_to_offset_year_cumulative[0]
    ws4["G7"] = vat_to_offset_year_cumulative[1]
    ws4["E37"] = "040"
    ws4["G37"] = vat_to_offset_total

    for sn in ("list05", "list06", "list07", "list08"):
        wb.create_sheet(sn)

    return wb


def build_vat_registry_wb(
    *,
    sales: Sequence[tuple[str, str | None, str, str, float, float]] = (),
    purchases: Sequence[tuple[str, str | None, str, str, float, float]] = (),
) -> WorkbookT:
    """Ilova — реестр счетов-фактур (Приложение №4 к Расчёту НДС, 10 листов).

    Tuples format (по обеим спискам): (counterparty_name, counterparty_inn,
    invoice_no, invoice_date_str DD.MM.YYYY, amount_excl_vat, vat_amount).

    - list01 (Таблица №1, **закупки**): шапка R11; данные R15+; колонки
      B seq | C name | D inn | E invoice_no | F invoice_date | G amount | H vat.
    - list02 (Таблица №2, **продажи**): аналогично + I=amount_incl_vat (=G+H).
    - list03..list10 — пустые скелеты остальных приложений.
    """
    wb = Workbook()
    ws1 = wb.active
    assert ws1 is not None
    ws1.title = "list01"
    # sentinel для format_detector
    ws1["B9"] = (
        "Реестр счетов-фактур (документов, заменяющих счета-фактуры) "
        "по приобретенным товарам (услугам) *"
    )
    ws1["B10"] = "Приложение № 4 таблица №1 к Расчету  налога на добавленную стоимость"
    ws1["B11"] = "№ "
    ws1["C11"] = "Наименование поставщика"
    ws1["D11"] = "ИНН поставщика **"
    ws1["E11"] = " счет-фактуры"
    ws1["G11"] = "Стоимость поставки\n(без НДС) ***"
    ws1["H11"] = "Сумма НДС"
    _write_registry_rows(ws1, purchases, with_total_with_vat=False)

    ws2 = wb.create_sheet("list02")
    ws2["B8"] = "Приложение № 4 таблица № 2 к Расчету налога на добавленную стоимость"
    ws2["B9"] = "Реестр счетов-фактур (документов, заменяющих счета-фактуры)"
    ws2["B10"] = "по реализованным товарам (услугам)"
    ws2["B11"] = "№ "
    ws2["C11"] = "Наименование покупателя "
    ws2["D11"] = "ИНН покупателя "
    ws2["E11"] = "счет-фактуры"
    ws2["G11"] = "Стоимость поставки \n(без НДС)"
    ws2["H11"] = "Сумма НДС"
    ws2["I11"] = "Стоимость с НДС"
    _write_registry_rows(ws2, sales, with_total_with_vat=True)

    for sn in ("list03", "list04", "list05", "list06", "list07", "list08", "list09", "list10"):
        wb.create_sheet(sn)

    return wb


def _write_registry_rows(
    ws: Any,
    rows: Sequence[tuple[str, str | None, str, str, float, float]],
    *,
    with_total_with_vat: bool,
) -> None:
    """Записать data rows реестра ЭСФ начиная с R15."""
    start_row = 15
    for idx, (name, inn, invoice_no, invoice_date, amount, vat) in enumerate(rows, start=1):
        r = start_row + idx - 1
        ws.cell(row=r, column=2, value=float(idx))  # B = seq
        ws.cell(row=r, column=3, value=name)  # C
        ws.cell(row=r, column=4, value=inn)  # D
        ws.cell(row=r, column=5, value=invoice_no)  # E
        ws.cell(row=r, column=6, value=invoice_date)  # F
        ws.cell(row=r, column=7, value=amount)  # G
        ws.cell(row=r, column=8, value=vat)  # H
        if with_total_with_vat:
            ws.cell(row=r, column=9, value=amount + vat)  # I


def build_form2_income_statement_wb(
    *,
    inn: float = 306399449,
    organization_name: str = '"AZ RUHDIL SAVDO" MCHJ',
    period_year: int = 2025,
    period_quarter: int = 4,
    # Все суммы в тыс. сум. (как в реальном файле). 'x' = неприменимо.
    # Defaults — реальные числа из form2_q4_2025_full.xltx (AZ RUHDIL SAVDO).
    # row 6 (010 Чистая выручка): D=prior, F=current. E/G всегда 'x'.
    revenue_prior: Any = 6559649.0,
    revenue_current: Any = 5973686.0,
    # row 7 (020 Себестоимость): E=prior expense, G=current expense.
    cost_prior: Any = 5820702.0,
    cost_current: Any = 5037111.0,
    # row 8 (030 Валовая прибыль): D=prior income, F=current income.
    gross_profit_prior: Any = 738947.0,
    gross_profit_current: Any = 936575.0,
    # row 15 (100 Прибыль от основной): (D, E) prior, (F, G) current.
    operating_profit_prior: tuple[Any, Any] = (27390.0, 0.0),
    operating_profit_current: tuple[Any, Any] = (87086.0, 0.0),
    # row 23 (180 Расходы в виде процентов): E=prior, G=current (expense only).
    interest_expense_prior: Any = 158407.0,
    interest_expense_current: Any = 40088.0,
    # row 29 (240 Прибыль до уплаты налога): (D, E) prior, (F, G) current.
    profit_before_tax_prior: tuple[Any, Any] = (0.0, 124633.0),
    profit_before_tax_current: tuple[Any, Any] = (63358.0, 0.0),
    # row 30 (250 Налог на прибыль): E=prior, G=current.
    profit_tax_prior: Any = 11389.0,
    profit_tax_current: Any = 19661.0,
    # row 32 (270 Чистая прибыль): (D, E) prior, (F, G) current. Signed via pair.
    net_profit_prior: tuple[Any, Any] = (0.0, 136022.0),
    net_profit_current: tuple[Any, Any] = (43697.0, 0.0),
) -> WorkbookT:
    """Форма №2 (Отчёт о финрезультатах). 3 листа.

    Координаты сверены с реальной выгрузкой Q4 2025 (form2_q4_2025_full.xltx,
    ИНН 306399449 AZ RUHDIL SAVDO). list02 — табличные данные, list03 пустой.

    Используй ``Any``-параметры с ``'x'`` или ``None``, чтобы прогнать парсер
    через best-effort cell-skipping ветки (CA-014).
    """
    wb = Workbook()
    ws1 = wb.active
    assert ws1 is not None
    ws1.title = "list01"
    ws1["B3"] = "Отчет о финансовых результатах - форма № 2"
    ws1["B5"] = "на"
    ws1["C5"] = float(period_year)
    ws1["D5"] = "год"
    ws1["E5"] = float(period_quarter)
    ws1["F5"] = "квартал"
    ws1["B8"] = "Предприяия, организация"
    ws1["C8"] = organization_name
    ws1["H18"] = "ИНН"
    ws1["I18"] = inn
    ws1["B24"] = "Единица измерения, тыс. сум."

    ws2 = wb.create_sheet("list02")
    # row 6 (010 Чистая выручка): E/G всегда 'x' (только income column)
    ws2["C6"] = "010"
    ws2["D6"] = revenue_prior
    ws2["E6"] = "x"
    ws2["F6"] = revenue_current
    ws2["G6"] = "x"
    # row 7 (020 Себестоимость): D/F всегда 'x' (только expense column)
    ws2["C7"] = "020"
    ws2["D7"] = "x"
    ws2["E7"] = cost_prior
    ws2["F7"] = "x"
    ws2["G7"] = cost_current
    # row 8 (030 Валовая прибыль): income only (E/G обычно 0 в реальном файле)
    ws2["C8"] = "030"
    ws2["D8"] = gross_profit_prior
    ws2["E8"] = 0.0
    ws2["F8"] = gross_profit_current
    ws2["G8"] = 0.0
    # row 15 (100 Прибыль от основной деятельности): signed pair
    ws2["C15"] = "100"
    ws2["D15"], ws2["E15"] = operating_profit_prior
    ws2["F15"], ws2["G15"] = operating_profit_current
    # row 23 (180 Расходы в виде процентов): D/F 'x', E/G — expense
    ws2["C23"] = "180"
    ws2["D23"] = "x"
    ws2["E23"] = interest_expense_prior
    ws2["F23"] = "x"
    ws2["G23"] = interest_expense_current
    # row 29 (240 Прибыль до уплаты налога): signed pair
    ws2["C29"] = "240"
    ws2["D29"], ws2["E29"] = profit_before_tax_prior
    ws2["F29"], ws2["G29"] = profit_before_tax_current
    # row 30 (250 Налог на прибыль): D/F 'x', E/G — expense
    ws2["C30"] = "250"
    ws2["D30"] = "x"
    ws2["E30"] = profit_tax_prior
    ws2["F30"] = "x"
    ws2["G30"] = profit_tax_current
    # row 32 (270 Чистая прибыль): signed pair
    ws2["C32"] = "270"
    ws2["D32"], ws2["E32"] = net_profit_prior
    ws2["F32"], ws2["G32"] = net_profit_current

    wb.create_sheet("list03")
    return wb


def build_form1_balance_sheet_wb(
    *,
    inn: float = 306399449,
    organization_name: str = '"AZ RUHDIL SAVDO" MCHJ',
    period_year: int = 2025,
    period_quarter: int = 4,
) -> WorkbookT:
    """Форма №1 (Бухбаланс). 4 листа. Минимум для format_detector."""
    wb = Workbook()
    ws1 = wb.active
    assert ws1 is not None
    ws1.title = "list01"
    ws1["B3"] = "Бухгалтерский баланс - форма № 1"
    ws1["B4"] = "на"
    ws1["C4"] = float(period_year)
    ws1["D4"] = "год"
    ws1["E4"] = float(period_quarter)
    ws1["F4"] = "квартал"
    ws1["B7"] = "Предприяия, организация"
    ws1["C7"] = organization_name
    ws1["H17"] = "ИНН"
    ws1["I17"] = inn
    for sn in ("list02", "list03", "list04"):
        wb.create_sheet(sn)
    return wb


def build_profit_tax_wb(
    *,
    inn: float = 306399449,
    organization_name: str = '"AZ RUHDIL SAVDO" MCHJ',
    period_year: int = 2026,
    period_quarter: int = 1,
) -> WorkbookT:
    """Расчёт налога на прибыль. 15 листов. Минимум для format_detector."""
    wb = Workbook()
    ws1 = wb.active
    assert ws1 is not None
    ws1.title = "list01"
    ws1["B2"] = "РАСЧЕТ\nналога на прибыль"
    ws1["B4"] = "ИНН"
    ws1["C4"] = inn
    ws1["C6"] = "Вид документа"
    ws1["D6"] = 1.0
    ws1["F6"] = "Отчётный период"
    ws1["G6"] = float(period_quarter)
    ws1["I6"] = "квартал"
    ws1["J6"] = "Налоговый период"
    ws1["K6"] = float(period_year)
    ws1["B8"] = "Полное наименование налогоплательщика"
    ws1["G8"] = organization_name
    for i in range(2, 16):
        wb.create_sheet(f"list{i:02d}")
    return wb
